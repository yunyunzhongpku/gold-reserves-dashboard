from __future__ import annotations

import argparse
import csv
import glob
from datetime import date, datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
MARKET_DIR = ROOT / "data" / "market"
WIND_DAILY_FILE = MARKET_DIR / "wind_daily.csv"

WIND_SERIES = {
    "gold_price": "SPTAUUSDOZ.IDC",
    "dollar_index": "USDX.FX",
    "gvz": "GVZ.GI",
}
START_DATE = "2003-01-01"
FIELDNAMES = ["date", "gold_price", "dollar_index", "gvz"]

# (sheet, 0-based value column) 与源 Excel 对账
EXCEL_ANCHORS = {
    "gold_price": ("实际利率与金价", 2),
    "dollar_index": ("美元指数", 1),
    "gvz": ("隐含波动率", 1),
}


def build_wind_daily_rows(series_by_key):
    all_dates = sorted({d for series in series_by_key.values() for d in series})
    rows = []
    for d in all_dates:
        row = {"date": d}
        for key in FIELDNAMES[1:]:
            row[key] = series_by_key.get(key, {}).get(d, "")
        rows.append(row)
    return rows


def compare_series(wind_series, excel_series, tol=1e-6):
    mismatches = []
    for d, wv in wind_series.items():
        ev = excel_series.get(d)
        if ev in (None, ""):
            continue
        ev = float(ev)
        if abs(wv - ev) > tol * max(1.0, abs(ev)):
            mismatches.append((d, wv, ev))
    return sorted(mismatches)


def fetch_wind_series(codes, begin, end):
    from tkf_wind import w
    w.start()
    out = {}
    for key, code in codes.items():
        data = w.wsd(code, "close", begin, end, "")
        if getattr(data, "ErrorCode", -1) != 0:
            raise RuntimeError(f"Wind wsd failed for {code}: ErrorCode={data.ErrorCode}")
        series = {}
        for t, v in zip(data.Times, data.Data[0]):
            if v is None:
                continue
            day = t.date() if isinstance(t, datetime) else t
            series[day.strftime("%Y-%m-%d")] = float(v)
        out[key] = series
    return out


def load_excel_anchor_series():
    from openpyxl import load_workbook
    path = glob.glob(str(ROOT / "data" / "*.xlsx"))[0]
    workbook = load_workbook(path, read_only=True, data_only=True)
    out = {}
    try:
        for key, (sheet, col) in EXCEL_ANCHORS.items():
            series = {}
            for raw in workbook[sheet].iter_rows(values_only=True):
                head = raw[0] if raw else None
                if isinstance(head, (date, datetime)) and len(raw) > col and isinstance(raw[col], (int, float)):
                    day = head.date() if isinstance(head, datetime) else head
                    series[day.strftime("%Y-%m-%d")] = float(raw[col])
            out[key] = series
    finally:
        workbook.close()
    return out


def write_csv(path, rows, fieldnames):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def refresh(begin=START_DATE, end=None, reconcile=True):
    end = end or date.today().strftime("%Y-%m-%d")
    series = fetch_wind_series(WIND_SERIES, begin, end)
    if reconcile:
        excel = load_excel_anchor_series()
        mismatches = {
            key: compare_series(series[key], excel.get(key, {}))
            for key in WIND_SERIES
        }
        failures = {key: found[:5] for key, found in mismatches.items() if found}
        if failures:
            raise SystemExit(f"Reconciliation failed vs Excel: {failures}")
    rows = build_wind_daily_rows(series)
    write_csv(WIND_DAILY_FILE, rows, FIELDNAMES)
    return rows


def main():
    parser = argparse.ArgumentParser(description="Refresh Wind daily series (LAN-only).")
    parser.add_argument("--begin", default=START_DATE)
    parser.add_argument("--end", default=None)
    parser.add_argument("--no-reconcile", action="store_true")
    args = parser.parse_args()
    rows = refresh(begin=args.begin, end=args.end, reconcile=not args.no_reconcile)
    print(f"Wrote {WIND_DAILY_FILE} ({len(rows)} rows)")


if __name__ == "__main__":
    main()
