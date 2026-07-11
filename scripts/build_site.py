from pathlib import Path
import calendar
import csv
import json
import math
from datetime import date, datetime
from html import escape

ROOT = Path(__file__).resolve().parents[1]
DATA_FILE = ROOT / "data" / "招商证券：黄金图表整理2607.xlsx"
MARKET_DIR = ROOT / "data" / "market"
BREAKEVEN_FILE = MARKET_DIR / "fred_t10yie.csv"
COT_FILE = MARKET_DIR / "cftc_gold_cot.csv"
OFFICIAL_RESERVES_MANUAL_FILE = MARKET_DIR / "official_reserves_manual.csv"
SITE_DIR = ROOT / "site"
OUTPUT_FILE = SITE_DIR / "index.html"

TEN_THOUSAND_TROY_OZ_TO_TONNES = 0.311034768
OFFICIAL_RESERVE_COLUMNS = [
    "date", "china_reserves_10k_oz", "global_reserves", "source"]

WIND_DAILY_FILE = MARKET_DIR / "wind_daily.csv"
DFII10_FILE = MARKET_DIR / "fred_dfii10.csv"
SHEET_REAL_RATE = "实际利率与金价"
SHEET_DOLLAR = "美元指数"
SHEET_VOLATILITY = "隐含波动率"
SHEET_ETF = "黄金ETF持有量&期现基差"
SHEET_RESERVES = "官方黄金储备"
SHEET_VALUATION = "中长期估值指标"
SHEET_EPU = "经济政策不确定性"
SHEET_GPR = "地缘政治风险"

STATE_LABELS = {
    "supportive": "支持",
    "neutral": "中性",
    "headwind": "压力",
    "planned": "待接入",
    "missing": "缺失",
}

QUALITY_LABELS = {
    "fresh": "新鲜",
    "stale": "滞后",
    "very-stale": "严重滞后",
    "planned": "待接入",
    "missing": "缺失",
    "partial": "部分可用",
}

FREQUENCY_THRESHOLDS = {
    "daily": (4, 14),
    "weekly": (10, 21),
    "monthly": (45, 75),
}

DRIVER_LAYER_IDS = (
    "real_rate",
    "dollar",
    "inflation_expectation",
    "official_reserves",
    "positioning_technical",
)

GOLD_CHART_RANGES = {"3m": 3, "1y": 12, "3y": 36}
GOLD_CHART_KEYS = ("gold_price", "ma20", "ma60", "ma200")

GOLD_CHART_SCRIPT = r"""(() => {
  const root = document.querySelector("[data-gold-chart]");
  if (!root) return;

  const payloadElement = root.querySelector("#gold-chart-data");
  if (!payloadElement) return;

  let payload;
  try {
    payload = JSON.parse(payloadElement.textContent);
  } catch (error) {
    return;
  }
  if (!payload || !payload.ranges) return;

  const rangeButtons = [...root.querySelectorAll("[data-chart-range]")];
  const seriesButtons = [...root.querySelectorAll("[data-chart-series]")];
  const charts = [...root.querySelectorAll(".gold-chart")];
  let activeRange = "1y";

  const setText = (selector, value, numeric = true) => {
    const element = root.querySelector(selector);
    if (!element) return;
    element.textContent = value == null
      ? "—"
      : numeric
        ? Number(value).toLocaleString("zh-CN", {maximumFractionDigits: 2})
        : value;
  };

  const showValues = (point) => {
    if (!point) return;
    setText("[data-tooltip-date]", point.date, false);
    setText("[data-tooltip-price]", point.gold_price);
    setText("[data-tooltip-ma20]", point.ma20);
    setText("[data-tooltip-ma60]", point.ma60);
    setText("[data-tooltip-ma200]", point.ma200);
  };

  const showPoint = (svg, event) => {
    const points = payload.ranges[activeRange];
    if (!points || !points.length) return;
    if (!svg || !event) return;

    const rect = svg.getBoundingClientRect();
    if (!rect.width) return;
    const viewWidth = svg.viewBox.baseVal.width || 720;
    const viewX = (event.clientX - rect.left) / rect.width * viewWidth;
    const firstX = points[0].x;
    const lastX = points[points.length - 1].x;
    const span = lastX - firstX || 1;
    const ratio = Math.max(0, Math.min(1, (viewX - firstX) / span));
    const point = points[Math.round(ratio * (points.length - 1))];
    if (!point) return;

    const line = svg.querySelector("[data-hover-line]");
    const marker = svg.querySelector("[data-hover-price]");
    if (line) {
      line.removeAttribute("hidden");
      line.setAttribute("x1", point.x);
      line.setAttribute("x2", point.x);
    }
    if (marker && point.gold_price_y != null) {
      marker.removeAttribute("hidden");
      marker.setAttribute("cx", point.x);
      marker.setAttribute("cy", point.gold_price_y);
    }
    showValues(point);
  };

  const activateRange = (rangeId) => {
    const points = payload.ranges[rangeId];
    if (!points || !points.length) return;
    activeRange = rangeId;
    rangeButtons.forEach((button) => {
      button.setAttribute(
        "aria-pressed",
        String(button.dataset.chartRange === activeRange)
      );
    });
    charts.forEach((svg) => {
      svg.toggleAttribute("hidden", svg.dataset.range !== activeRange);
    });
    showValues(points[points.length - 1]);
  };

  rangeButtons.forEach((button) => {
    button.addEventListener("click", () => activateRange(button.dataset.chartRange));
  });

  seriesButtons.forEach((button) => {
    button.addEventListener("click", () => {
      if (button.disabled) return;
      const next = button.getAttribute("aria-pressed") !== "true";
      button.setAttribute("aria-pressed", String(next));
      root.classList.toggle(`hide-${button.dataset.chartSeries}`, !next);
    });
  });

  charts.forEach((svg) => {
    svg.addEventListener("pointermove", (event) => showPoint(svg, event));
    svg.addEventListener("pointerdown", (event) => showPoint(svg, event));
  });

  activateRange("1y");
})();"""


def classify_staleness(latest_date, today, frequency):
    if not latest_date:
        return "missing", None
    latest = datetime.strptime(latest_date, "%Y-%m-%d").date()
    lag = max(0, (today - latest).days)   # 月末前瞻日期(如储备盖到月末)不显示负滞后
    fresh_max, stale_max = FREQUENCY_THRESHOLDS[frequency]
    if lag <= fresh_max:
        return "fresh", lag
    if lag <= stale_max:
        return "stale", lag
    return "very-stale", lag


def format_date(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if value is None:
        return ""
    return str(value)[:10]


def as_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def is_number(value):
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def fmt_number(value, digits=2):
    if value is None:
        return "—"
    return f"{value:.{digits}f}"


def fmt_integer(value):
    if value is None:
        return "—"
    return f"{value:,.0f}"


def fmt_signed(value, digits=2, suffix=""):
    if value is None:
        return "—"
    return f"{value:+.{digits}f}{suffix}"


def fmt_signed_integer(value, suffix=""):
    if value is None:
        return "—"
    return f"{value:+,.0f}{suffix}"


def load_workbook():
    try:
        from openpyxl import load_workbook as open_workbook
    except ImportError as exc:
        raise RuntimeError("读取 Excel 数据需要 openpyxl，请在当前 Python 环境中安装后重试。") from exc

    if not DATA_FILE.exists():
        raise FileNotFoundError(f"未找到数据文件：{DATA_FILE}")

    return open_workbook(DATA_FILE, read_only=True, data_only=True)


def read_sheet_rows(workbook, sheet_name, columns, max_date=None):
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"未找到工作表：{sheet_name}")

    sheet = workbook[sheet_name]
    rows = []
    for raw in sheet.iter_rows(values_only=True):
        row_date = as_date(raw[0] if raw else None)
        if row_date is None:
            continue
        if max_date is not None and row_date > max_date:
            continue

        item = {"date": format_date(row_date), "_date": row_date}
        has_value = False
        for key, col_index in columns.items():
            value = raw[col_index] if len(raw) > col_index else None
            if is_number(value):
                item[key] = float(value)
                has_value = True
            else:
                item[key] = None

        if has_value:
            rows.append(item)

    rows.sort(key=lambda row: row["_date"])
    return rows


def read_csv_rows(path, numeric_columns):
    if not path.exists():
        return []

    rows = []
    with path.open(encoding="utf-8", newline="") as file:
        for raw in csv.DictReader(file):
            row_date = as_date(datetime.strptime(raw["date"], "%Y-%m-%d"))
            item = {"date": format_date(row_date), "_date": row_date}
            has_value = False
            for key in numeric_columns:
                value = raw.get(key, "")
                if value == "":
                    item[key] = None
                    continue
                item[key] = float(value)
                has_value = True
            for key, value in raw.items():
                if key not in item and key not in numeric_columns:
                    item[key] = value
            if has_value:
                rows.append(item)

    rows.sort(key=lambda row: row["_date"])
    return rows


def month_index(value):
    return value.year * 12 + value.month


def optional_csv_float(value):
    return None if value in (None, "") else float(value)


def read_official_reserve_rows(path, max_date=None):
    if not path.exists():
        return []

    rows = []
    seen_months = set()
    previous_date = None
    previous_month = None
    with path.open(encoding="utf-8", newline="") as file:
        reader = csv.DictReader(file)
        if reader.fieldnames != OFFICIAL_RESERVE_COLUMNS:
            expected = ",".join(OFFICIAL_RESERVE_COLUMNS)
            actual = ",".join(reader.fieldnames or [])
            raise ValueError(
                f"官方储备 CSV 表头必须严格为 {expected}；实际为 {actual}")

        for raw in reader:
            row_date = datetime.strptime(raw["date"], "%Y-%m-%d").date()
            current_month = month_index(row_date)
            if current_month in seen_months:
                raise ValueError(f"官方储备月份重复：{raw['date']}")
            if previous_date is not None and row_date <= previous_date:
                raise ValueError("官方储备日期必须严格升序")
            if previous_month is not None and current_month != previous_month + 1:
                raise ValueError("官方储备月份必须连续")

            ounces = float(raw["china_reserves_10k_oz"])
            if not math.isfinite(ounces) or ounces <= 0:
                raise ValueError("中国官方黄金储备必须为有限正数")
            source = (raw.get("source") or "").strip()
            if not source:
                raise ValueError("中国官方黄金储备来源不能为空")
            global_reserves = optional_csv_float(raw.get("global_reserves"))
            if global_reserves is not None and (not math.isfinite(global_reserves) or global_reserves <= 0):
                raise ValueError("全球官方黄金储备必须为有限正数")
            rows.append({
                "date": format_date(row_date),
                "_date": row_date,
                "china_reserves_10k_oz": ounces,
                "china_reserves": ounces * TEN_THOUSAND_TROY_OZ_TO_TONNES,
                "global_reserves": global_reserves,
                "china_source": source,
                "global_source": source if global_reserves is not None else None,
            })
            seen_months.add(current_month)
            previous_date = row_date
            previous_month = current_month

    return [row for row in rows if max_date is None or row["_date"] <= max_date]


def merge_reserve_rows(base_rows, override_rows):
    merged = {}
    for row in base_rows:
        item = {
            **row,
            "china_reserves_10k_oz": row.get("china_reserves_10k_oz"),
            "china_source": row.get("china_source") or "Excel: 官方黄金储备",
            "global_source": row.get("global_source") or "Excel: 官方黄金储备",
        }
        merged[(row["_date"].year, row["_date"].month)] = item

    for row in override_rows:
        key = (row["_date"].year, row["_date"].month)
        item = merged.get(key, {
            "date": row["date"], "_date": row["_date"],
            "china_reserves": None, "global_reserves": None,
            "china_source": None, "global_source": None,
        })
        item["china_reserves_10k_oz"] = row["china_reserves_10k_oz"]
        item["china_reserves"] = row["china_reserves"]
        item["china_source"] = row["china_source"]
        if row["global_reserves"] is not None:
            item["global_reserves"] = row["global_reserves"]
            item["global_source"] = row["global_source"]
        merged[key] = item

    return sorted(merged.values(), key=lambda row: row["_date"])


def latest_with_value(rows, key):
    for row in reversed(rows):
        if row.get(key) is not None:
            return row
    return None


def merge_override_rows(base_rows, override_rows):
    if not override_rows:
        return base_rows

    first_override_date = override_rows[0]["_date"]
    rows = [row for row in base_rows if row["_date"] < first_override_date]
    rows.extend(override_rows)
    rows.sort(key=lambda row: row["_date"])
    return rows


def optional_change(current, previous, key):
    if current is None or previous is None:
        return None
    if current.get(key) is None or previous.get(key) is None:
        return None
    return current[key] - previous[key]


def change_over_observations(rows, key, lookback):
    valid = [row for row in rows if row.get(key) is not None]
    if len(valid) < 2:
        return None, None

    latest = valid[-1]
    prior_index = max(0, len(valid) - lookback - 1)
    prior = valid[prior_index]
    return latest[key] - prior[key], prior["date"]


def state_from_delta(delta, supportive_when="down", threshold=0.0):
    if delta is None or abs(delta) <= threshold:
        return "neutral"
    if supportive_when == "down":
        return "supportive" if delta < 0 else "headwind"
    return "supportive" if delta > 0 else "headwind"


def percentile_rank(rows, key, lookback=None):
    valid = [row for row in rows if row.get(key) is not None]
    if not valid:
        return None
    if lookback:
        valid = valid[-lookback:]
    if len(valid) < 2:
        return None

    latest_value = valid[-1][key]
    below_or_equal = sum(1 for row in valid if row[key] <= latest_value)
    return below_or_equal / len(valid)


def pct_return(current, previous):
    if current is None or previous in (None, 0):
        return None
    return current / previous - 1


def trailing_return(rows, key, lookback):
    valid = [row[key] for row in rows if row.get(key) is not None]
    if len(valid) <= lookback:
        return None
    return pct_return(valid[-1], valid[-1 - lookback])


def attach_rolling_ma(rows, key, window, out_key):
    prices = []
    output = []
    for row in rows:
        value = row.get(key)
        if value is None:
            output.append({**row, out_key: None})
            continue
        prices.append(value)
        ma = sum(prices[-window:]) / window if len(prices) >= window else None
        output.append({**row, out_key: ma})
    return output


def attach_gold_technicals(rows):
    output = attach_rolling_ma(rows, "gold_price", 20, "ma20")
    output = attach_rolling_ma(output, "gold_price", 60, "ma60")
    return attach_rolling_ma(output, "gold_price", 200, "ma200")


def classify_technical_state(price, ma20, ma60, ma200, return_5d):
    if ma20 is None or return_5d is None:
        short_term = "样本不足"
    elif price > ma20 and return_5d > 0:
        short_term = "短期反弹"
    elif price < ma20 and return_5d < 0:
        short_term = "短期走弱"
    else:
        short_term = "短期震荡"

    if ma60 is None or ma200 is None:
        medium_term = "样本不足"
    elif price > ma60 and price > ma200:
        medium_term = "中期偏多"
    elif price < ma60 and price < ma200:
        medium_term = "中期偏空"
    else:
        medium_term = "中期修复"

    if ma20 is None or ma60 is None or ma200 is None:
        alignment = "样本不足"
    elif price >= ma20 > ma60 > ma200:
        alignment = "多头排列"
    elif price <= ma20 < ma60 < ma200:
        alignment = "空头排列"
    else:
        alignment = "未形成完整排列"

    return {
        "short_term": short_term,
        "medium_term": medium_term,
        "alignment": alignment,
    }


def correlation(values):
    pairs = [(x, y) for x, y in values if x is not None and y is not None]
    if len(pairs) < 6:
        return None

    xs = [x for x, _ in pairs]
    ys = [y for _, y in pairs]
    mean_x = sum(xs) / len(xs)
    mean_y = sum(ys) / len(ys)
    cov = sum((x - mean_x) * (y - mean_y) for x, y in pairs)
    var_x = sum((x - mean_x) ** 2 for x in xs)
    var_y = sum((y - mean_y) ** 2 for y in ys)
    if var_x == 0 or var_y == 0:
        return None
    return cov / (var_x * var_y) ** 0.5


def fmt_corr(value):
    if value is None:
        return "—"
    return f"{value:+.2f}"


def fmt_pct(value):
    if value is None:
        return "—"
    return f"{value * 100:+.1f}%"


def fmt_axis_value(value):
    if value is None:
        return "—"
    if abs(value) >= 1000:
        return f"{value:,.0f}"
    if abs(value) >= 100:
        return f"{value:.0f}"
    if abs(value) >= 10:
        return f"{value:.1f}"
    return f"{value:.2f}"


def corr_tone(value, expected="negative"):
    if value is None:
        return "样本不足"
    if expected == "absolute":
        if abs(value) >= 0.35:
            return "有效"
        if abs(value) <= 0.1:
            return "失效"
        return "偏弱"
    if expected == "negative":
        if value <= -0.35:
            return "有效"
        if value >= 0.1:
            return "失效"
        return "偏弱"
    if value >= 0.35:
        return "有效"
    if value <= -0.1:
        return "失效"
    return "偏弱"


def corr_class(value, expected="negative"):
    tone = corr_tone(value, expected=expected)
    if tone == "有效":
        return "supportive"
    if tone == "失效":
        return "headwind"
    return "neutral"


def phase_label(start, end):
    if end is None:
        return f"{start.year}-至今"
    if start.year == end.year:
        return str(start.year)
    return f"{start.year}-{end.year}"


def filter_date_range(rows, start, end):
    return [
        row for row in rows
        if row.get("_date") is not None and row["_date"] >= start and (end is None or row["_date"] <= end)
    ]


def build_change_pairs(rows, factor_key, gold_key, lookback):
    pairs = []
    for i in range(lookback, len(rows)):
        current = rows[i]
        previous = rows[i - lookback]
        factor_current = current.get(factor_key)
        factor_previous = previous.get(factor_key)
        gold_current = current.get(gold_key)
        gold_previous = previous.get(gold_key)
        if factor_current is None or factor_previous is None or gold_current is None or gold_previous is None:
            continue
        pairs.append({
            "date": current["date"],
            "_date": current["_date"],
            "factor_change": factor_current - factor_previous,
            "gold_return": pct_return(gold_current, gold_previous),
        })
    return pairs


def rolling_corr(rows, x_key, y_key, window):
    output = []
    for i in range(window - 1, len(rows)):
        sample = rows[i - window + 1:i + 1]
        corr = correlation((row.get(x_key), row.get(y_key)) for row in sample)
        output.append({
            "date": rows[i]["date"],
            "_date": rows[i]["_date"],
            "corr": corr,
        })
    return [row for row in output if row["corr"] is not None]


def corr_for_range(rows, x_key, y_key, start, end):
    sample = filter_date_range(rows, start, end)
    return correlation((row.get(x_key), row.get(y_key)) for row in sample), len(sample)


def gold_price_asof(gold_rows, target_date):
    price = None
    for row in gold_rows:
        if row["_date"] > target_date:
            break
        if row.get("gold_price") is not None:
            price = row["gold_price"]
    return price


def build_reserve_relationship_rows(reserve_rows, gold_rows):
    rows = []
    changes = []
    for i, row in enumerate(reserve_rows):
        if i == 0:
            china_change = None
            global_change = None
        else:
            china_change = optional_change(row, reserve_rows[i - 1], "china_reserves")
            global_change = optional_change(row, reserve_rows[i - 1], "global_reserves")
        changes.append((china_change, global_change))

        if i < 3:
            continue
        china_changes = [change[0] for change in changes[i - 2:i + 1] if change[0] is not None]
        global_changes = [change[1] for change in changes[i - 2:i + 1] if change[1] is not None]
        china_3m = sum(china_changes) if china_changes else None
        global_3m = sum(global_changes) if global_changes else None
        gold_current = gold_price_asof(gold_rows, row["_date"])
        gold_previous = gold_price_asof(gold_rows, reserve_rows[i - 3]["_date"])
        gold_return_3m = pct_return(gold_current, gold_previous)
        if china_3m is None or gold_return_3m is None:
            continue
        rows.append({
            "date": row["date"],
            "_date": row["_date"],
            "china_3m_change": china_3m,
            "global_3m_change": global_3m,
            "gold_price": gold_current,
            "gold_return_3m": gold_return_3m,
        })
    return rows


def latest_significant_cycle_start(rows, key, threshold, min_months=3):
    runs = []
    run_start = None
    run_rows = []
    for row in rows:
        if row.get(key) is not None and row[key] >= threshold:
            if run_start is None:
                run_start = row
                run_rows = []
            run_rows.append(row)
        elif run_start is not None:
            if len(run_rows) >= min_months:
                runs.append(run_rows)
            run_start = None
            run_rows = []
    if run_start is not None and len(run_rows) >= min_months:
        runs.append(run_rows)
    if not runs:
        return None
    return runs[-1][0]["date"]


def weak_periods_from_rolling(rows, threshold=0.1, max_periods=3):
    periods = []
    active = []
    for row in rows:
        corr = row.get("corr")
        weak = corr is not None and abs(corr) <= threshold
        if weak:
            active.append(row)
        elif active:
            if len(active) >= 3:
                periods.append((active[0]["date"], active[-1]["date"]))
            active = []
    if len(active) >= 3:
        periods.append((active[0]["date"], active[-1]["date"]))
    return periods[-max_periods:]


def phase_gold_return(gold_rows, start, end):
    sample = filter_date_range([row for row in gold_rows if row.get("gold_price") is not None], start, end)
    if len(sample) < 2:
        return None
    return pct_return(sample[-1]["gold_price"], sample[0]["gold_price"])


def attach_gold_price(rows, gold_rows):
    attached = []
    for row in rows:
        gold_price = row.get("gold_price")
        if gold_price is None:
            gold_price = gold_price_asof(gold_rows, row["_date"])
        if gold_price is None:
            continue
        attached.append({**row, "gold_price": gold_price})
    return attached


def quality_badge(quality):
    return QUALITY_LABELS.get(quality, quality)


def state_badge(state):
    return STATE_LABELS.get(state, state)


def score_layers(layers):
    weights = {"supportive": 1, "headwind": -1, "neutral": 0}
    scored = [weights[layer["state"]] for layer in layers if layer["state"] in weights]
    active = len(scored)
    score = sum(scored)
    tendency = score / active if active else 0.0
    if tendency >= 0.25:
        posture, posture_state = "偏多", "supportive"
    elif tendency <= -0.25:
        posture, posture_state = "承压", "headwind"
    else:
        posture, posture_state = "中性", "neutral"
    return score, posture, posture_state, tendency, active


def combine_subsignal_states(states):
    weights = {"supportive": 1, "neutral": 0, "headwind": -1}
    votes = [weights[state] for state in states if state in weights]
    if not votes:
        return "missing"
    average = sum(votes) / len(votes)
    if average >= 0.5:
        return "supportive"
    if average <= -0.5:
        return "headwind"
    return "neutral"


def score_driver_layers(layers):
    selected = [
        layer for layer in layers
        if layer.get("id") in DRIVER_LAYER_IDS
        and layer.get("state") in {"supportive", "neutral", "headwind"}
        and layer.get("data_quality") not in {"missing", "very-stale"}
    ]
    if len(selected) < 3:
        return 0, "数据不足", "missing", 0.0, len(selected)
    return score_layers(selected)


def make_real_rate_layer(rows):
    latest = latest_with_value(rows, "real_rate")
    delta, since = change_over_observations(rows, "real_rate", 21)
    state = state_from_delta(delta, supportive_when="down", threshold=0.05)

    return {
        "id": "real_rate",
        "name": "实际利率",
        "source": "FRED: DFII10 (10Y TIPS real yield)",
        "frequency": "daily",
        "data_quality": "fresh",
        "state": state,
        "latest": {
            "date": latest["date"],
            "value": latest["real_rate"],
            "gold_price": latest.get("gold_price"),
        },
        "change": delta,
        "change_since": since,
        "value_label": f"{fmt_number(latest['real_rate'])}%",
        "change_label": fmt_signed(delta, suffix="pct"),
        "read": (
            f"10Y 实际利率最新 {fmt_number(latest['real_rate'])}%，"
            f"近约 1 个月变化 {fmt_signed(delta, suffix='pct')}。"
        ),
        "wrong_if": "实际利率继续上行并维持高位，黄金机会成本压力扩大。",
        "next_trigger": "更新 10Y TIPS real yield，并观察与金价的滚动相关是否重新增强。",
        "chart_key": "real_rate",
        "chart_rows": rows,
    }


def make_dollar_layer(rows):
    latest = latest_with_value(rows, "dollar_index")
    delta, since = change_over_observations(rows, "dollar_index", 21)
    state = state_from_delta(delta, supportive_when="down", threshold=0.5)

    return {
        "id": "dollar",
        "name": "美元",
        "source": "Wind: USDX.FX",
        "frequency": "daily",
        "data_quality": "fresh",
        "state": state,
        "latest": {
            "date": latest["date"],
            "value": latest["dollar_index"],
            "gold_price": latest.get("gold_price"),
        },
        "change": delta,
        "change_since": since,
        "value_label": fmt_number(latest["dollar_index"]),
        "change_label": fmt_signed(delta),
        "read": (
            f"美元指数最新 {fmt_number(latest['dollar_index'])}，"
            f"近约 1 个月变化 {fmt_signed(delta)}。"
        ),
        "wrong_if": "美元重新走强，且实际利率同步上行，美元计价黄金承压。",
        "next_trigger": "更新美元指数，并与实际利率层一起判断是否形成同向压力。",
        "chart_key": "dollar_index",
        "chart_rows": rows,
    }


def make_missing_inflation_layer():
    return {
        "id": "inflation_expectation",
        "name": "通胀预期",
        "source": "FRED: T10YIE (10-Year Breakeven Inflation Rate)",
        "frequency": "daily",
        "data_quality": "missing",
        "state": "missing",
        "latest": {"date": None, "value": None},
        "change": None,
        "change_since": None,
        "value_label": "缺数据",
        "change_label": "—",
        "read": "当前未找到本地 FRED T10YIE 数据，暂不对通胀预期层打分。",
        "wrong_if": "通胀预期快速回落而实际利率上行，黄金的通胀保护逻辑会减弱。",
        "next_trigger": "刷新 FRED T10YIE 后，观察通胀预期与实际利率的方向差。",
        "chart_key": None,
        "chart_rows": [],
    }


def make_inflation_layer(rows):
    latest = latest_with_value(rows, "breakeven_10y")
    if latest is None:
        return make_missing_inflation_layer()

    delta, since = change_over_observations(rows, "breakeven_10y", 21)
    state = state_from_delta(delta, supportive_when="up", threshold=0.05)

    return {
        "id": "inflation_expectation",
        "name": "通胀预期",
        "source": "FRED: T10YIE (10-Year Breakeven Inflation Rate)",
        "frequency": "daily",
        "data_quality": "fresh",
        "state": state,
        "latest": {
            "date": latest["date"],
            "value": latest["breakeven_10y"],
        },
        "change": delta,
        "change_since": since,
        "value_label": f"{fmt_number(latest['breakeven_10y'])}%",
        "change_label": fmt_signed(delta, suffix="pct"),
        "read": (
            f"10Y breakeven 最新 {fmt_number(latest['breakeven_10y'])}%，"
            f"近约 1 个月变化 {fmt_signed(delta, suffix='pct')}。"
        ),
        "wrong_if": "通胀预期快速回落而实际利率上行，黄金的通胀保护逻辑会减弱。",
        "next_trigger": "观察 10Y breakeven 与实际利率是否同向上行，区分通胀补偿和真实利率压力。",
        "chart_key": "breakeven_10y",
        "chart_rows": rows,
    }


def make_reserve_layer(rows):
    latest = rows[-1]
    previous = rows[-2] if len(rows) >= 2 else None
    china_change = optional_change(latest, previous, "china_reserves")
    global_change = optional_change(latest, previous, "global_reserves")

    state = state_from_delta(china_change, supportive_when="up", threshold=0.0)

    chart_rows = []
    for i, row in enumerate(rows):
        if i == 0:
            china_mom_change = None
            global_mom_change = None
        else:
            china_mom_change = optional_change(row, rows[i - 1], "china_reserves")
            global_mom_change = optional_change(row, rows[i - 1], "global_reserves")
        chart_rows.append({
            **row,
            "china_mom_change": china_mom_change,
            "global_mom_change": global_mom_change,
        })

    if latest.get("global_reserves") is None or global_change is None:
        global_read = "全球官方黄金储备本期未更新。"
    else:
        global_read = (
            f"全球官方黄金储备 {fmt_number(latest['global_reserves'])} 吨，"
            f"环比 {fmt_signed(global_change, suffix=' 吨')}。"
        )

    return {
        "id": "official_reserves",
        "name": "央行购金",
        "source": latest.get("china_source") or "Excel: 官方黄金储备",
        "frequency": "monthly",
        "data_quality": "fresh",
        "state": state,
        "latest": {
            "date": latest["date"],
            "value": latest["china_reserves"],
            "china_reserves": latest["china_reserves"],
            "global_reserves": latest["global_reserves"],
            "china_change": china_change,
            "global_change": global_change,
            "china_source": latest.get("china_source") or "Excel: 官方黄金储备",
            "global_source": (
                latest.get("global_source")
                if latest.get("global_reserves") is not None else None),
        },
        "change": china_change,
        "change_since": previous["date"] if previous else None,
        "value_label": f"中国 {fmt_number(latest['china_reserves'])} 吨",
        "change_label": fmt_signed(china_change, suffix=" 吨"),
        "read": (
            f"中国官方黄金储备 {fmt_number(latest['china_reserves'])} 吨，"
            f"环比 {fmt_signed(china_change, suffix=' 吨')}；"
            f"{global_read}"
        ),
        "wrong_if": "中国与全球官方储备同时转为持续净卖出，慢变量支撑减弱。",
        "next_trigger": "等待下一期官方储备数据，重点看中国与全球是否同向增加。",
        "chart_key": "china_mom_change",
        "chart_type": "bar",
        "chart_label": "央行购金环比柱状图",
        "chart_rows": chart_rows,
    }


def make_positioning_layer(etf_rows, vol_rows, gold_rows, cot_rows, today=None):
    today = today or date.today()
    for row in etf_rows:
        if row.get("spdr_holdings") is not None and row.get("ishares_holdings") is not None:
            row["etf_total"] = row["spdr_holdings"] + row["ishares_holdings"]
        else:
            row["etf_total"] = None

    latest_etf = latest_with_value(etf_rows, "etf_total")
    etf_delta, etf_since = change_over_observations(etf_rows, "etf_total", 21)
    latest_vol = latest_with_value(vol_rows, "gvz")
    latest_gold = latest_with_value(gold_rows, "gold_price")
    gold_delta, _ = change_over_observations(gold_rows, "gold_price", 60)

    etf_date = latest_etf["date"] if latest_etf else None
    etf_quality, etf_lag = classify_staleness(etf_date, today, "daily")
    gvz_date = latest_vol["date"] if latest_vol else None
    gvz_quality, gvz_lag = classify_staleness(gvz_date, today, "daily")

    etf_state = (
        state_from_delta(etf_delta, supportive_when="up", threshold=3.0)
        if latest_etf else "missing"
    )
    if etf_quality == "very-stale":
        etf_state = "missing"

    latest_cot = latest_with_value(cot_rows, "managed_money_net")
    cot_delta, _ = change_over_observations(cot_rows, "managed_money_net", 4)
    cot_percentile = percentile_rank(cot_rows, "managed_money_net", 156)
    cot_date = latest_cot["date"] if latest_cot else None
    cot_quality, cot_lag = classify_staleness(cot_date, today, "weekly")
    if latest_cot is None:
        cot_state = "missing"
        cot_sentence = "CFTC Managed Money 数据缺失，仓位拥挤度暂不计入。"
    else:
        if cot_percentile is not None and cot_percentile > 0.9:
            cot_state = "headwind"
        elif cot_delta is not None and cot_delta < -15000:
            cot_state = "headwind"
        elif cot_delta is not None and cot_delta > 15000:
            cot_state = "supportive"
        else:
            cot_state = "neutral"
        cot_sentence = (
            f"Managed Money 黄金净多 {fmt_integer(latest_cot['managed_money_net'])} 张，"
            f"近 4 周变化 {fmt_signed_integer(cot_delta, suffix=' 张')}，"
            f"三年分位 {fmt_number((cot_percentile or 0) * 100)}%。"
        )
    if cot_quality == "very-stale":
        cot_state = "missing"

    state = combine_subsignal_states([etf_state, cot_state])
    active_qualities = [
        (quality, lag)
        for signal_state, quality, lag in [
            (etf_state, etf_quality, etf_lag),
            (cot_state, cot_quality, cot_lag),
        ]
        if signal_state != "missing"
    ]
    if any(quality == "fresh" for quality, _ in active_qualities):
        data_quality = "fresh"
    elif any(quality == "stale" for quality, _ in active_qualities):
        data_quality = "stale"
    else:
        data_quality = "missing"
    active_lags = [lag for _, lag in active_qualities if lag is not None]
    lag_days = min(active_lags) if active_lags else None
    latest_dates = [value for value in [etf_date, cot_date] if value]
    latest_date = max(latest_dates) if latest_dates else None

    gvz_percentile = percentile_rank(vol_rows, "gvz", 756) if latest_vol else None
    return {
        "id": "positioning_technical",
        "name": "仓位与技术",
        "source": "Wind: ETF(SPDR/iShares)、GVZ；CFTC: Disaggregated COT Gold",
        "frequency": "daily",
        "data_quality": data_quality,
        "lag_days": lag_days,
        "state": state,
        "sub_states": {"etf": etf_state, "cot": cot_state},
        "latest": {
            "date": latest_date,
            "value": latest_etf["etf_total"] if latest_etf else None,
            "etf_total": latest_etf["etf_total"] if latest_etf else None,
            "spdr_holdings": latest_etf["spdr_holdings"] if latest_etf else None,
            "ishares_holdings": latest_etf["ishares_holdings"] if latest_etf else None,
            "etf_change": etf_delta,
            "etf_date": etf_date,
            "etf_quality": etf_quality,
            "etf_lag": etf_lag,
            "gvz": latest_vol["gvz"] if latest_vol else None,
            "gvz_percentile": gvz_percentile,
            "gvz_date": gvz_date,
            "gvz_quality": gvz_quality,
            "gvz_lag": gvz_lag,
            "gold_price": latest_gold["gold_price"] if latest_gold else None,
            "gold_change_60": gold_delta,
            "cot_date": cot_date,
            "cot_quality": cot_quality,
            "cot_lag": cot_lag,
            "managed_money_net": latest_cot["managed_money_net"] if latest_cot else None,
            "managed_money_long": latest_cot["managed_money_long"] if latest_cot else None,
            "managed_money_short": latest_cot["managed_money_short"] if latest_cot else None,
            "managed_money_net_change": cot_delta,
            "managed_money_net_percentile": cot_percentile,
        },
        "change": etf_delta,
        "change_since": etf_since,
        "value_label": f"ETF {fmt_number(latest_etf['etf_total'] if latest_etf else None)} 吨",
        "change_label": fmt_signed(etf_delta, suffix=" 吨"),
        "read": (
            f"SPDR+iShares 黄金 ETF 合计 {fmt_number(latest_etf['etf_total'] if latest_etf else None)} 吨，"
            f"近约 1 个月变化 {fmt_signed(etf_delta, suffix=' 吨')}；"
            f"GVZ {fmt_number(latest_vol['gvz'] if latest_vol else None)}，"
            f"三年分位 {fmt_number((gvz_percentile or 0) * 100)}%。"
            f"{cot_sentence}"
            f"GVZ 数据 {gvz_date or '—'}({QUALITY_LABELS.get(gvz_quality, gvz_quality)},滞后 {gvz_lag if gvz_lag is not None else '—'} 天)。"
        ),
        "wrong_if": "ETF 持有量持续流出、价格跌破中期趋势，且 Managed Money 净多回落或拥挤度过高。",
        "next_trigger": "每周更新 CFTC COT，观察 ETF 流量与 Managed Money 净多是否同向。",
        "chart_key": "etf_total",
        "chart_rows": etf_rows,
    }


def driver_row(
    row_id, name, category, state, value, change, read,
    quality, date_value, source,
):
    return {
        "id": row_id,
        "name": name,
        "category": category,
        "state": state,
        "value": value,
        "change": change,
        "read": read,
        "quality": quality,
        "date": date_value,
        "source": source,
    }


def make_driver_rows(layers):
    by_id = {layer["id"]: layer for layer in layers}
    real = by_id["real_rate"]
    dollar = by_id["dollar"]
    inflation = by_id["inflation_expectation"]
    reserve = by_id["official_reserves"]
    positioning = by_id["positioning_technical"]
    latest_positioning = positioning["latest"]

    return [
        driver_row(
            "real_rate", "实际利率", "宏观", real["state"],
            real["value_label"], real["change_label"], "机会成本变化",
            real["data_quality"], real["latest"].get("date"), real["source"],
        ),
        driver_row(
            "dollar", "美元", "宏观", dollar["state"],
            dollar["value_label"], dollar["change_label"], "美元计价压力",
            dollar["data_quality"], dollar["latest"].get("date"), dollar["source"],
        ),
        driver_row(
            "official_reserves", "中国央行购金", "官方部门", reserve["state"],
            fmt_signed(reserve["latest"].get("china_change"), suffix=" 吨"),
            "较上月增量", "官方购金慢变量", reserve["data_quality"],
            reserve["latest"].get("date"),
            reserve["latest"].get("china_source") or reserve["source"],
        ),
        driver_row(
            "inflation_expectation", "通胀预期", "宏观", inflation["state"],
            inflation["value_label"], inflation["change_label"], "通胀补偿变化",
            inflation["data_quality"], inflation["latest"].get("date"),
            inflation["source"],
        ),
        driver_row(
            "etf", "黄金 ETF", "资金流", positioning["sub_states"]["etf"],
            f"{fmt_number(latest_positioning.get('etf_total'))} 吨",
            fmt_signed(latest_positioning.get("etf_change"), suffix=" 吨"),
            "黄金 ETF 合计持仓", latest_positioning["etf_quality"],
            latest_positioning.get("etf_date"), "Wind: ETF(SPDR/iShares)",
        ),
        driver_row(
            "cot", "CFTC 净多", "仓位", positioning["sub_states"]["cot"],
            f"{fmt_integer(latest_positioning.get('managed_money_net'))} 张",
            fmt_signed_integer(
                latest_positioning.get("managed_money_net_change"), suffix=" 张"),
            "Managed Money 近 4 周变化", latest_positioning["cot_quality"],
            latest_positioning.get("cot_date"), "CFTC: Disaggregated COT Gold",
        ),
    ]


def make_recent_changes(layers):
    by_id = {layer["id"]: layer for layer in layers}
    real = by_id["real_rate"]
    dollar = by_id["dollar"]
    reserve = by_id["official_reserves"]
    positioning = by_id["positioning_technical"]

    macro_states = {real["state"], dollar["state"]}
    macro_quality_missing = any(
        layer.get("data_quality") in {"missing", "very-stale"}
        for layer in [real, dollar]
    )
    if macro_quality_missing or not macro_states <= {"supportive", "neutral", "headwind"}:
        macro_tone, macro_headline = "missing", "实际利率或美元数据不完整"
    elif real["state"] == dollar["state"] == "neutral":
        macro_tone, macro_headline = "neutral", "实际利率与美元均偏稳"
    elif real["state"] == dollar["state"] == "headwind":
        macro_tone, macro_headline = "headwind", "实际利率与美元同向上行"
    elif real["state"] == dollar["state"] == "supportive":
        macro_tone, macro_headline = "supportive", "实际利率与美元同向回落"
    else:
        macro_tone, macro_headline = "neutral", "实际利率与美元信号分化"

    chart_rows = reserve.get("chart_rows", [])
    current_purchase = reserve["latest"].get("china_change")
    previous_purchase = (
        chart_rows[-2].get("china_mom_change") if len(chart_rows) >= 2 else None)
    if current_purchase is None:
        official_tone, official_headline = "missing", "中国央行购金数据缺失"
    elif current_purchase == 0:
        official_tone, official_headline = "neutral", "中国央行本月未继续增持"
    elif current_purchase < 0:
        official_tone, official_headline = "headwind", "中国央行转为净卖出"
    elif previous_purchase is not None and current_purchase > previous_purchase:
        official_tone, official_headline = "supportive", "中国央行购金加速"
    elif previous_purchase is not None and current_purchase < previous_purchase:
        official_tone = "supportive"
        official_headline = "中国央行仍在净买入，但速度放缓"
    else:
        official_tone, official_headline = "supportive", "中国央行继续净买入"

    etf_state = positioning["sub_states"]["etf"]
    cot_state = positioning["sub_states"]["cot"]
    valid_states = {"supportive", "neutral", "headwind"}
    if etf_state not in valid_states or cot_state not in valid_states:
        flows_tone, flows_headline = "missing", "ETF 或 CFTC 数据不完整"
    elif etf_state == cot_state:
        flows_tone = etf_state
        flows_headline = (
            "ETF 与 CFTC 仓位均为中性"
            if etf_state == "neutral" else "ETF 与 CFTC 仓位同向"
        )
    else:
        flows_tone, flows_headline = "neutral", "ETF 与 CFTC 仓位分化"

    return [
        {
            "id": "macro",
            "tone": macro_tone,
            "label": "宏观条件",
            "headline": macro_headline,
            "detail": f"实际利率 {real['change_label']}；美元 {dollar['change_label']}。",
        },
        {
            "id": "official",
            "tone": official_tone,
            "label": "官方购金",
            "headline": official_headline,
            "detail": (
                f"本月 {fmt_signed(current_purchase, suffix=' 吨')}；"
                f"上月 {fmt_signed(previous_purchase, suffix=' 吨')}。"
            ),
        },
        {
            "id": "flows",
            "tone": flows_tone,
            "label": "资金与仓位",
            "headline": flows_headline,
            "detail": (
                f"ETF {fmt_signed(positioning['latest'].get('etf_change'), suffix=' 吨')}；"
                "CFTC "
                f"{fmt_signed_integer(positioning['latest'].get('managed_money_net_change'), suffix=' 张')}。"
            ),
        },
    ]


def make_price_trend_layer(gold_rows):
    chart_rows = attach_gold_technicals(gold_rows)
    latest = latest_with_value(chart_rows, "gold_price")
    return_5d = trailing_return(chart_rows, "gold_price", 5)
    momentum_3m = trailing_return(chart_rows, "gold_price", 63)
    prices = [row["gold_price"] for row in chart_rows if row.get("gold_price") is not None]
    peak = max(prices[-252:]) if prices else None
    drawdown = pct_return(latest["gold_price"], peak) if peak else None
    technical = classify_technical_state(
        latest["gold_price"],
        latest.get("ma20"),
        latest.get("ma60"),
        latest.get("ma200"),
        return_5d,
    )
    if technical["medium_term"] == "中期偏空":
        trigger = (
            f"重新站上 MA60 {fmt_number(latest.get('ma60'))} 与 "
            f"MA200 {fmt_number(latest.get('ma200'))} 后，转为中期偏多。"
        )
    elif technical["medium_term"] == "中期偏多":
        trigger = (
            f"同时跌破 MA60 {fmt_number(latest.get('ma60'))} 与 "
            f"MA200 {fmt_number(latest.get('ma200'))} 后，中期趋势转弱。"
        )
    else:
        trigger = "等待价格同时站上或跌破 MA60 与 MA200，确认中期方向。"
    technical["trigger"] = trigger
    state = {
        "中期偏多": "supportive",
        "中期偏空": "headwind",
    }.get(technical["medium_term"], "neutral")

    return {
        "id": "price_trend",
        "name": "价格与趋势",
        "source": "Wind: SPTAUUSDOZ.IDC（派生 MA20/MA60/MA200、动量）",
        "frequency": "daily",
        "data_quality": "fresh",
        "state": state,
        "latest": {
            "date": latest["date"],
            "value": latest["gold_price"],
            "gold_price": latest["gold_price"],
            "ma20": latest.get("ma20"),
            "ma60": latest.get("ma60"),
            "ma200": latest.get("ma200"),
            "gap_ma20": pct_return(latest["gold_price"], latest.get("ma20")),
            "gap_ma60": pct_return(latest["gold_price"], latest.get("ma60")),
            "gap_ma200": pct_return(latest["gold_price"], latest.get("ma200")),
            "return_5d": return_5d,
            "momentum_3m": momentum_3m,
            "drawdown": drawdown,
        },
        "technical": technical,
        "change": momentum_3m,
        "change_since": None,
        "value_label": fmt_number(latest["gold_price"]),
        "change_label": fmt_pct(momentum_3m),
        "read": (
            f"{technical['medium_term']}，{technical['short_term']}，"
            f"{technical['alignment']}。"
        ),
        "wrong_if": "价格重新站上或跌破中长期均线后，技术状态将按同一规则更新。",
        "next_trigger": "跟踪价格与 MA20、MA60、MA200 的相对位置。",
        "chart_key": "gold_price",
        "chart_rows": chart_rows,
    }


def make_monthly_state_layer(layer_id, name, source, rows, key, gold_rows, unit=""):
    latest = latest_with_value(rows, key)
    delta, since = change_over_observations(rows, key, 3)        # 近3个月变化
    threshold = abs(latest[key]) * 0.05 if latest else 0.0       # 5% 显著阈值,抗噪
    state = state_from_delta(delta, supportive_when="up", threshold=threshold)
    return {
        "id": layer_id, "name": name, "source": source, "frequency": "monthly",
        "data_quality": "fresh", "state": state,
        "latest": {"date": latest["date"], "value": latest[key]},
        "change": delta, "change_since": since,
        "value_label": f"{fmt_number(latest[key])}{unit}",
        "change_label": fmt_signed(delta, suffix=unit),
        "read": (
            f"{name} 最新 {fmt_number(latest[key])}{unit},近3个月 {fmt_signed(delta, suffix=unit)}。"
            f"作为风险偏好/避险代理观察,非稳定因果驱动,详见关系检验。"
        ),
        "wrong_if": f"{name} 回落且金价同向走弱,避险溢价消退。",
        "next_trigger": f"更新 {name},观察其与金价的滚动相关是否有效。",
        "chart_key": key, "chart_rows": rows,
    }


def make_epu_layer(epu_rows, gold_rows):
    return make_monthly_state_layer(
        "epu", "经济政策不确定性", "Excel: 经济政策不确定性(下轮转 Wind)", epu_rows, "epu", gold_rows)


def make_gpr_layer(gpr_rows, gold_rows):
    return make_monthly_state_layer(
        "gpr", "地缘政治风险", "Excel: 地缘政治风险(下轮转 Wind)", gpr_rows, "gpr", gold_rows)


def make_valuation_snapshot(rows):
    latest = rows[-1] if rows else None
    if latest is None:
        return None
    return {
        "date": latest["date"],
        "gold_price": latest.get("gold_price"),
        "gold_to_m2": latest.get("gold_to_m2"),
        "valuation_percentile": latest.get("valuation_percentile"),
    }


def make_monthly_factor_relationship(rel_id, name, metric, rows, factor_key, gold_rows, expected="positive"):
    paired = attach_gold_price([r for r in rows if r.get(factor_key) is not None], gold_rows)
    enriched = []
    for i in range(3, len(paired)):
        factor_3m = paired[i][factor_key] - paired[i - 3][factor_key]
        gold_3m = pct_return(paired[i]["gold_price"], paired[i - 3]["gold_price"])
        if gold_3m is None:
            continue
        enriched.append({
            "date": paired[i]["date"], "_date": paired[i]["_date"],
            "factor_3m_change": factor_3m, "gold_return_3m": gold_3m,
            factor_key: paired[i][factor_key], "gold_price": paired[i]["gold_price"],
        })
    rolling = rolling_corr(enriched, "factor_3m_change", "gold_return_3m", 24)
    latest_corr = rolling[-1]["corr"] if rolling else None

    phase_defs = [
        ("2018-2020", date(2018, 1, 1), date(2020, 12, 31)),
        ("2021-2022", date(2021, 1, 1), date(2022, 12, 31)),
        ("2023-2024H1", date(2023, 1, 1), date(2024, 6, 30)),
        ("2024H2-至今", date(2024, 7, 1), None),
    ]
    phases = []
    for label, start, end in phase_defs:
        corr, obs = corr_for_range(enriched, "factor_3m_change", "gold_return_3m", start, end)
        phases.append({"label": label, "corr": corr, "tone": corr_tone(corr, expected=expected),
                       "observations": obs, "gold_return": phase_gold_return(gold_rows, start, end)})

    latest_row = enriched[-1] if enriched else {}
    return {
        "id": rel_id, "name": name, "metric": metric, "expected": expected,
        "latest_corr": latest_corr, "latest_tone": corr_tone(latest_corr, expected=expected),
        "start_month": None, "weak_periods": weak_periods_from_rolling(rolling),
        "trend_rows": enriched, "factor_key": factor_key, "factor_label": name,
        "gold_key": "gold_price", "rolling_corr": rolling, "phases": phases,
        "latest": {"date": latest_row.get("date"),
                   "factor_change": latest_row.get("factor_3m_change"),
                   "gold_return": latest_row.get("gold_return_3m")},
        "read": (
            f"{name} 与金价的 3个月滚动相关 {fmt_corr(latest_corr)};"
            f"作为避险/风险偏好代理,关系常间歇有效,勿当稳定因果。"
        ),
    }


def make_central_bank_relationship(reserve_rows, gold_rows):
    relationship_rows = build_reserve_relationship_rows(reserve_rows, gold_rows)
    rolling = rolling_corr(relationship_rows, "china_3m_change", "gold_return_3m", 24)
    latest_corr = rolling[-1]["corr"] if rolling else None
    start_month = latest_significant_cycle_start(
        relationship_rows,
        "china_3m_change",
        threshold=30,
        min_months=3,
    )

    phases = []
    phase_defs = [
        ("2018-2020", date(2018, 1, 1), date(2020, 12, 31)),
        ("2021-2022", date(2021, 1, 1), date(2022, 12, 31)),
        ("2023-2024H1", date(2023, 1, 1), date(2024, 6, 30)),
        ("2024H2-至今", date(2024, 7, 1), None),
    ]
    for label, start, end in phase_defs:
        corr, observations = corr_for_range(relationship_rows, "china_3m_change", "gold_return_3m", start, end)
        phases.append({
            "label": label,
            "corr": corr,
            "tone": corr_tone(corr, expected="positive"),
            "observations": observations,
            "gold_return": phase_gold_return(gold_rows, start, end),
        })

    latest_row = relationship_rows[-1] if relationship_rows else {}
    return {
        "id": "central_bank_purchases",
        "name": "央行购金",
        "metric": "中国官方黄金储备 3个月增量 vs 黄金 3个月收益",
        "expected": "positive",
        "latest_corr": latest_corr,
        "latest_tone": corr_tone(latest_corr, expected="positive"),
        "start_month": start_month,
        "weak_periods": weak_periods_from_rolling(rolling),
        "trend_rows": relationship_rows,
        "factor_key": "china_3m_change",
        "factor_label": "购金3个月增量",
        "gold_key": "gold_price",
        "rolling_corr": rolling,
        "phases": phases,
        "latest": {
            "date": latest_row.get("date"),
            "factor_change": latest_row.get("china_3m_change"),
            "gold_return": latest_row.get("gold_return_3m"),
        },
        "read": (
            f"什么时候开始显著推动：按中国官方黄金储备 3 个月增量连续超过 30 吨识别，"
            f"最近一轮从 {start_month or '—'} 开始。"
        ),
    }


def make_real_rate_relationship(real_rate_rows):
    return make_two_horizon_relationship(
        "real_rate",
        "实际利率",
        "实际利率变化 vs 黄金收益",
        real_rate_rows,
        "real_rate",
        expected="negative",
        short_label="实际利率短期",
        medium_label="实际利率中期",
        read_suffix="负值越强，实际利率对黄金的解释越有效。",
    )


def make_two_horizon_relationship(
    relationship_id,
    name,
    metric,
    rows,
    factor_key,
    expected,
    short_label,
    medium_label,
    read_suffix,
):
    short_pairs = build_change_pairs(rows, factor_key, "gold_price", 21)
    medium_pairs = build_change_pairs(rows, factor_key, "gold_price", 63)
    short_rolling = rolling_corr(short_pairs, "factor_change", "gold_return", 126)
    medium_rolling = rolling_corr(medium_pairs, "factor_change", "gold_return", 252)
    short_latest = short_rolling[-1]["corr"] if short_rolling else None
    medium_latest = medium_rolling[-1]["corr"] if medium_rolling else None

    phase_defs = [
        ("2003-2011", date(2003, 1, 1), date(2011, 12, 31)),
        ("2012-2018", date(2012, 1, 1), date(2018, 12, 31)),
        ("2019-2021", date(2019, 1, 1), date(2021, 12, 31)),
        ("2022-2024", date(2022, 1, 1), date(2024, 12, 31)),
        ("2025-至今", date(2025, 1, 1), None),
    ]
    phases = []
    for label, start, end in phase_defs:
        short_corr, short_obs = corr_for_range(short_pairs, "factor_change", "gold_return", start, end)
        medium_corr, medium_obs = corr_for_range(medium_pairs, "factor_change", "gold_return", start, end)
        phases.append({
            "label": label,
            "short_corr": short_corr,
            "medium_corr": medium_corr,
            "short_tone": corr_tone(short_corr, expected=expected),
            "medium_tone": corr_tone(medium_corr, expected=expected),
            "observations": min(short_obs, medium_obs),
            "gold_return": phase_gold_return(rows, start, end),
        })

    return {
        "id": relationship_id,
        "name": name,
        "metric": metric,
        "expected": expected,
        "short_term": {
            "label": short_label,
            "window": "21日变化 vs 21日黄金收益，126日滚动相关",
            "latest_corr": short_latest,
            "tone": corr_tone(short_latest, expected=expected),
            "rolling_corr": short_rolling,
        },
        "medium_term": {
            "label": medium_label,
            "window": "63日变化 vs 63日黄金收益，252日滚动相关",
            "latest_corr": medium_latest,
            "tone": corr_tone(medium_latest, expected=expected),
            "rolling_corr": medium_rolling,
        },
        "latest_corr": medium_latest,
        "latest_tone": corr_tone(medium_latest, expected=expected),
        "rolling_corr": medium_rolling,
        "trend_rows": [row for row in rows if row.get(factor_key) is not None and row.get("gold_price") is not None],
        "factor_key": factor_key,
        "factor_label": name,
        "gold_key": "gold_price",
        "phases": phases,
        "weak_periods": weak_periods_from_rolling(medium_rolling),
        "read": (
            f"短期滚动相关 {fmt_corr(short_latest)}，中期滚动相关 {fmt_corr(medium_latest)}；"
            f"{read_suffix}"
        ),
    }


def build_trend_forward_pairs(gold_rows, ma_window, horizon):
    points = [
        (row["date"], row["_date"], row["gold_price"])
        for row in gold_rows if row.get("gold_price") is not None
    ]
    pairs = []
    for i in range(ma_window - 1, len(points) - horizon):
        ma = sum(p[2] for p in points[i - ma_window + 1:i + 1]) / ma_window
        signal = points[i][2] / ma - 1                       # t 时点趋势缺口
        forward = points[i + horizon][2] / points[i][2] - 1   # 未来 horizon 收益
        pairs.append({"date": points[i][0], "_date": points[i][1],
                      "factor_change": signal, "gold_return": forward})
    return pairs


def make_trend_relationship(gold_rows):
    short_pairs = build_trend_forward_pairs(gold_rows, 200, 21)
    medium_pairs = build_trend_forward_pairs(gold_rows, 200, 63)
    short_rolling = rolling_corr(short_pairs, "factor_change", "gold_return", 126)
    medium_rolling = rolling_corr(medium_pairs, "factor_change", "gold_return", 252)
    short_latest = short_rolling[-1]["corr"] if short_rolling else None
    medium_latest = medium_rolling[-1]["corr"] if medium_rolling else None

    phase_defs = [
        ("2012-2018", date(2012, 1, 1), date(2018, 12, 31)),
        ("2019-2021", date(2019, 1, 1), date(2021, 12, 31)),
        ("2022-2024", date(2022, 1, 1), date(2024, 12, 31)),
        ("2025-至今", date(2025, 1, 1), None),
    ]
    phases = []
    for label, start, end in phase_defs:
        short_corr, short_obs = corr_for_range(short_pairs, "factor_change", "gold_return", start, end)
        medium_corr, medium_obs = corr_for_range(medium_pairs, "factor_change", "gold_return", start, end)
        phases.append({
            "label": label, "short_corr": short_corr, "medium_corr": medium_corr,
            "short_tone": corr_tone(short_corr, expected="positive"),
            "medium_tone": corr_tone(medium_corr, expected="positive"),
            "observations": min(short_obs, medium_obs),
            "gold_return": phase_gold_return(gold_rows, start, end),
        })

    return {
        "id": "price_trend", "name": "价格与趋势",
        "metric": "200日趋势缺口(t) vs 未来黄金收益(前瞻,防自证)", "expected": "positive",
        "short_term": {"label": "趋势→未来1月", "window": "200日缺口 vs 未来21日收益,126日滚动",
                       "latest_corr": short_latest, "tone": corr_tone(short_latest, expected="positive"),
                       "rolling_corr": short_rolling},
        "medium_term": {"label": "趋势→未来3月", "window": "200日缺口 vs 未来63日收益,252日滚动",
                        "latest_corr": medium_latest, "tone": corr_tone(medium_latest, expected="positive"),
                        "rolling_corr": medium_rolling},
        "latest_corr": medium_latest, "latest_tone": corr_tone(medium_latest, expected="positive"),
        "rolling_corr": medium_rolling,
        "trend_rows": attach_rolling_ma(
            [row for row in gold_rows if row.get("gold_price") is not None], "gold_price", 200, "ma200"),
        "factor_key": "ma200", "factor_label": "200日均线", "gold_key": "gold_price",
        "phases": phases, "weak_periods": weak_periods_from_rolling(medium_rolling),
        "read": (
            f"前瞻相关 短 {fmt_corr(short_latest)} / 中 {fmt_corr(medium_latest)};"
            f"正值表示趋势对未来收益有跟随效应。检验用未来收益,避免与同期金价自证。"
        ),
    }


def make_dollar_relationship(dollar_rows):
    return make_two_horizon_relationship(
        "dollar",
        "美元",
        "美元指数变化 vs 黄金收益",
        dollar_rows,
        "dollar_index",
        expected="negative",
        short_label="美元短期",
        medium_label="美元中期",
        read_suffix="负值越强，美元走强对黄金的压制越稳定。",
    )


def make_inflation_relationship(breakeven_rows, gold_rows):
    rows = attach_gold_price(breakeven_rows, gold_rows)
    return make_two_horizon_relationship(
        "inflation_expectation",
        "通胀预期",
        "10Y breakeven 变化 vs 黄金收益",
        rows,
        "breakeven_10y",
        expected="positive",
        short_label="通胀预期短期",
        medium_label="通胀预期中期",
        read_suffix="正值越强，通胀补偿对黄金的解释越稳定。",
    )


def make_positioning_sub_metric(metric_id, name, rows, factor_key, expected, gold_rows=None):
    relationship_rows = attach_gold_price(rows, gold_rows) if gold_rows else rows
    pairs = build_change_pairs(relationship_rows, factor_key, "gold_price", 21)
    rolling = rolling_corr(pairs, "factor_change", "gold_return", 52)
    latest_corr = rolling[-1]["corr"] if rolling else None
    phase_defs = [
        ("2021-2022", date(2021, 1, 1), date(2022, 12, 31)),
        ("2023-2024", date(2023, 1, 1), date(2024, 12, 31)),
        ("2025-至今", date(2025, 1, 1), None),
    ]
    phases = []
    for label, start, end in phase_defs:
        corr, observations = corr_for_range(pairs, "factor_change", "gold_return", start, end)
        phases.append({
            "label": label,
            "corr": corr,
            "tone": corr_tone(corr, expected=expected),
            "observations": observations,
            "gold_return": phase_gold_return(relationship_rows, start, end),
        })
    return {
        "id": metric_id,
        "name": name,
        "expected": expected,
        "latest_corr": latest_corr,
        "latest_tone": corr_tone(latest_corr, expected=expected),
        "rolling_corr": rolling,
        "trend_rows": [
            row for row in relationship_rows
            if row.get(factor_key) is not None and row.get("gold_price") is not None
        ],
        "factor_key": factor_key,
        "factor_label": name,
        "gold_key": "gold_price",
        "phases": phases,
        "weak_periods": weak_periods_from_rolling(rolling),
    }


def make_positioning_relationship(etf_rows, cot_rows, vol_rows, gold_rows):
    for row in etf_rows:
        if row.get("spdr_holdings") is not None and row.get("ishares_holdings") is not None:
            row["etf_total"] = row["spdr_holdings"] + row["ishares_holdings"]
    sub_metrics = [
        make_positioning_sub_metric(
            "etf_holdings",
            "ETF 持仓",
            etf_rows,
            "etf_total",
            expected="positive",
            gold_rows=gold_rows,
        ),
        make_positioning_sub_metric(
            "managed_money",
            "Managed Money 净多",
            cot_rows,
            "managed_money_net",
            expected="positive",
            gold_rows=gold_rows,
        ),
        make_positioning_sub_metric(
            "gvz",
            "GVZ 波动率",
            vol_rows,
            "gvz",
            expected="absolute",
        ),
    ]
    latest_corrs = [item["latest_corr"] for item in sub_metrics if item["latest_corr"] is not None]
    latest_corr = sum(latest_corrs) / len(latest_corrs) if latest_corrs else None
    return {
        "id": "positioning_technical",
        "name": "仓位与技术",
        "metric": "ETF、CFTC Managed Money、GVZ 与黄金收益",
        "expected": "mixed",
        "latest_corr": latest_corr,
        "latest_tone": "分项观察",
        "sub_metrics": sub_metrics,
        "read": "仓位与技术层拆成 ETF 持仓、Managed Money 净多和 GVZ 波动率，分别看其对黄金短期收益的解释力。",
    }


def make_relationships(
    reserve_rows, real_rate_rows, dollar_rows, breakeven_rows,
    etf_rows, cot_rows, vol_rows, gold_rows, epu_rows, gpr_rows,
):
    return [
        make_central_bank_relationship(reserve_rows, gold_rows),
        make_real_rate_relationship(real_rate_rows),
        make_dollar_relationship(dollar_rows),
        make_inflation_relationship(breakeven_rows, gold_rows),
        make_positioning_relationship(etf_rows, cot_rows, vol_rows, gold_rows),
        make_trend_relationship(gold_rows),
        make_monthly_factor_relationship("epu", "经济政策不确定性",
            "EPU 3个月变化 vs 黄金3个月收益", epu_rows, "epu", gold_rows, expected="positive"),
        make_monthly_factor_relationship("gpr", "地缘政治风险",
            "GPR 3个月变化 vs 黄金3个月收益", gpr_rows, "gpr", gold_rows, expected="positive"),
    ]


def read_dashboard_data(today=None):
    today = today or date.today()
    workbook = load_workbook()
    try:
        reserve_rows = read_sheet_rows(workbook, SHEET_RESERVES, {"china_reserves": 1, "global_reserves": 2}, max_date=today)
        etf_rows = read_sheet_rows(workbook, SHEET_ETF, {"spdr_holdings": 1, "ishares_holdings": 2}, max_date=today)
        valuation_rows = read_sheet_rows(workbook, SHEET_VALUATION, {"gold_price": 7, "gold_to_m2": 9, "valuation_percentile": 10}, max_date=today)
        epu_rows = read_sheet_rows(workbook, SHEET_EPU, {"epu": 1}, max_date=today)
        gpr_rows = read_sheet_rows(workbook, SHEET_GPR, {"gpr": 1}, max_date=today)
    finally:
        workbook.close()

    manual_reserve_rows = read_official_reserve_rows(
        OFFICIAL_RESERVES_MANUAL_FILE, max_date=today)
    reserve_rows = merge_reserve_rows(reserve_rows, manual_reserve_rows)
    source_files = [str(DATA_FILE.relative_to(ROOT))]
    if manual_reserve_rows:
        source_files.append(str(OFFICIAL_RESERVES_MANUAL_FILE.relative_to(ROOT)))

    wind_rows = read_csv_rows(WIND_DAILY_FILE, ["gold_price", "dollar_index", "gvz"])
    # 稀疏宽表 → 派生每指标的稠密子集,使行号 lookback == 有效观测 lookback
    gold_rows = [row for row in wind_rows if row.get("gold_price") is not None]
    dollar_rows = [row for row in wind_rows if row.get("dollar_index") is not None]
    vol_rows = [row for row in wind_rows if row.get("gvz") is not None]
    real_rate_rows = attach_gold_price(read_csv_rows(DFII10_FILE, ["real_rate"]), gold_rows)
    breakeven_rows = read_csv_rows(BREAKEVEN_FILE, ["breakeven_10y"])
    cot_rows = read_csv_rows(COT_FILE, [
        "open_interest", "managed_money_long", "managed_money_short",
        "managed_money_spread", "managed_money_net", "managed_money_net_to_oi",
    ])

    layers = [
        make_real_rate_layer(real_rate_rows),
        make_dollar_layer(dollar_rows),
        make_inflation_layer(breakeven_rows),
        make_reserve_layer(reserve_rows),
        make_positioning_layer(etf_rows, vol_rows, gold_rows, cot_rows, today),
        make_price_trend_layer(gold_rows),
        make_epu_layer(epu_rows, gold_rows),
        make_gpr_layer(gpr_rows, gold_rows),
    ]

    for layer in layers:
        if layer["id"] == "positioning_technical":
            continue
        quality, lag = classify_staleness(layer["latest"].get("date"), today, layer["frequency"])
        layer["data_quality"] = quality
        layer["lag_days"] = lag

    driver_layers = [layer for layer in layers if layer["id"] in DRIVER_LAYER_IDS]
    score, posture, posture_state, tendency, active_layers = score_driver_layers(
        driver_layers)
    technical_layer = next(layer for layer in layers if layer["id"] == "price_trend")
    driver_rows = make_driver_rows(layers)
    recent_changes = make_recent_changes(layers)

    return {
        "title": "黄金数据驱动跟踪",
        "source_file": " + ".join(source_files),
        "score": score, "tendency": tendency, "active_layers": active_layers,
        "posture": posture, "posture_state": posture_state,
        "layers": layers,
        "driver_layers": driver_layers,
        "driver_rows": driver_rows,
        "recent_changes": recent_changes,
        "technical_layer": technical_layer,
        "valuation": make_valuation_snapshot(valuation_rows),
        "reserve_rows": layers[3]["chart_rows"],
        "relationships": make_relationships(
            reserve_rows, real_rate_rows, dollar_rows, breakeven_rows,
            etf_rows, cot_rows, vol_rows, gold_rows, epu_rows, gpr_rows),
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


def shift_months(value: date, months: int) -> date:
    month_zero = value.year * 12 + value.month - 1 + months
    year, month_zero = divmod(month_zero, 12)
    month = month_zero + 1
    day = min(value.day, calendar.monthrange(year, month)[1])
    return date(year, month, day)


def build_gold_chart_ranges(rows: list[dict]) -> dict[str, list[dict]]:
    ranges = {range_id: [] for range_id in GOLD_CHART_RANGES}
    valid_rows = sorted(
        (
            row
            for row in rows
            if row.get("_date") is not None and row.get("gold_price") is not None
        ),
        key=lambda row: row["_date"],
    )
    if not valid_rows:
        return ranges

    latest_date = valid_rows[-1]["_date"]
    return {
        range_id: [
            row
            for row in valid_rows
            if row["_date"] >= shift_months(latest_date, -months)
        ]
        for range_id, months in GOLD_CHART_RANGES.items()
    }


def make_gold_chart_geometry(
    rows: list[dict], width: int = 720, height: int = 280
) -> list[dict]:
    if not rows:
        return []

    left, right, top, bottom = 54, 18, 18, 34
    plot_width = width - left - right
    plot_height = height - top - bottom
    values = [
        row[key]
        for row in rows
        for key in GOLD_CHART_KEYS
        if row.get(key) is not None
    ]
    low = min(values) if values else 0.0
    high = max(values) if values else 0.0
    span = high - low or 1.0
    denominator = max(1, len(rows) - 1)
    points = []
    for index, row in enumerate(rows):
        point = {
            "date": row["date"],
            "x": round(left + index / denominator * plot_width, 2),
        }
        for key in GOLD_CHART_KEYS:
            value = row.get(key)
            point[key] = value
            point[f"{key}_y"] = (
                None
                if value is None
                else round(top + (high - value) / span * plot_height, 2)
            )
        points.append(point)
    return points


def svg_series_path(points: list[dict], key: str) -> str:
    commands = []
    drawing = False
    for point in points:
        y = point.get(f"{key}_y")
        if y is None:
            drawing = False
            continue
        commands.append(f"{'L' if drawing else 'M'}{point['x']:.2f},{y:.2f}")
        drawing = True
    return " ".join(commands)


def make_gold_chart_svg(
    points: list[dict],
    range_id: str,
    hidden: bool = False,
    width: int = 720,
    height: int = 280,
) -> str:
    hidden_attr = " hidden" if hidden else ""
    range_label = {"3m": "3个月", "1y": "1年", "3y": "3年"}.get(
        range_id, range_id)
    first_date = points[0]["date"] if points else "—"
    last_date = points[-1]["date"] if points else "—"
    aria_label = (
        f"黄金价格与移动均线，{range_label}，美元每盎司，"
        f"{first_date}至{last_date}"
    )
    if not points:
        return f"""
    <svg class="gold-chart" data-range="{escape(range_id)}" viewBox="0 0 {width} {height}"
         role="img" aria-label="{escape(aria_label)}"{hidden_attr}>
      <text x="{width / 2:.0f}" y="{height / 2:.0f}" text-anchor="middle" class="chart-empty">暂无价格数据</text>
    </svg>
        """

    left, right, top, bottom = 54, 18, 18, 34
    values = [
        point[key]
        for point in points
        for key in GOLD_CHART_KEYS
        if point.get(key) is not None
    ]
    low = min(values)
    high = max(values)
    paths = "".join(
        f'<path class="chart-series series-{key}" '
        f'd="{svg_series_path(points, key)}"></path>'
        for key in GOLD_CHART_KEYS
    )
    return f"""
    <svg class="gold-chart" data-range="{escape(range_id)}" viewBox="0 0 {width} {height}"
         role="img" aria-label="{escape(aria_label)}"{hidden_attr}>
      <title>{escape(aria_label)}</title>
      <line x1="{left}" y1="{top}" x2="{left}" y2="{height - bottom}" class="y-axis"></line>
      <line x1="{left}" y1="{height - bottom}" x2="{width - right}" y2="{height - bottom}" class="x-axis"></line>
      <text x="4" y="{top + 4}" class="chart-label">{escape(fmt_axis_value(high))}</text>
      <text x="4" y="{height - bottom}" class="chart-label">{escape(fmt_axis_value(low))}</text>
      {paths}
      <line data-hover-line x1="{left}" y1="{top}" x2="{left}" y2="{height - bottom}" class="hover-line" hidden></line>
      <circle data-hover-price cx="{left}" cy="{top}" r="4" class="hover-price" hidden></circle>
      <text x="{left}" y="{height - 8}" class="chart-label">{escape(first_date)}</text>
      <text x="{width - right}" y="{height - 8}" text-anchor="end" class="chart-label">{escape(last_date)}</text>
    </svg>
    """


def safe_json(value) -> str:
    return json.dumps(
        value, ensure_ascii=False, separators=(",", ":")
    ).replace("<", "\\u003c")


def make_gold_chart_panel(technical_layer: dict) -> str:
    chart_rows = technical_layer.get("chart_rows") or []
    ranges = build_gold_chart_ranges(chart_rows)
    geometries = {
        range_id: make_gold_chart_geometry(rows)
        for range_id, rows in ranges.items()
    }
    charts = "".join(
        make_gold_chart_svg(
            geometries[range_id], range_id, hidden=(range_id != "1y"))
        for range_id in GOLD_CHART_RANGES
    )
    payload = safe_json({"ranges": geometries})
    latest = technical_layer.get("latest") or (
        chart_rows[-1] if chart_rows else {})

    def series_button(key: str, label: str) -> str:
        available = latest.get(key) is not None
        pressed = "true" if available else "false"
        disabled = "" if available else " disabled"
        return (
            f'<button type="button" data-chart-series="{key}" '
            f'aria-pressed="{pressed}"{disabled}>{escape(label)}</button>'
        )

    return f"""
    <div class="gold-chart-panel" data-gold-chart>
      <div class="chart-heading">
        <strong>金价与移动均线</strong>
        <span>美元/盎司 · 日期范围见横轴</span>
      </div>
      <div class="chart-controls" aria-label="黄金图表控制">
        <div class="range-controls" aria-label="日期范围">
          <button type="button" data-chart-range="3m" aria-pressed="false">3个月</button>
          <button type="button" data-chart-range="1y" aria-pressed="true">1年</button>
          <button type="button" data-chart-range="3y" aria-pressed="false">3年</button>
        </div>
        <div class="series-controls" aria-label="移动均线">
          {series_button("ma20", "20日均线")}
          {series_button("ma60", "60日均线")}
          {series_button("ma200", "200日均线")}
        </div>
      </div>
      <div class="chart-stack">{charts}</div>
      <div class="chart-tooltip" aria-live="polite">
        <strong data-tooltip-date>—</strong>
        <span>金价 <b data-tooltip-price>—</b></span>
        <span>MA20 <b data-tooltip-ma20>—</b></span>
        <span>MA60 <b data-tooltip-ma60>—</b></span>
        <span>MA200 <b data-tooltip-ma200>—</b></span>
      </div>
      <script type="application/json" id="gold-chart-data">{payload}</script>
    </div>
    """


def make_sparkline(rows, key, color="#2563eb", width=360, height=96, limit=120):
    points = [(row["date"], row.get(key)) for row in rows if row.get(key) is not None][-limit:]
    if len(points) < 2:
        return '<div class="empty-chart">暂无足够数据</div>'

    values = [value for _, value in points]
    min_value = min(values)
    max_value = max(values)
    span = max_value - min_value or 1
    left_pad = 42
    right_pad = 10
    top_pad = 12
    bottom_pad = 22
    chart_width = width - left_pad - right_pad
    chart_height = height - top_pad - bottom_pad

    coords = []
    for i, (_, value) in enumerate(points):
        x = left_pad + chart_width * i / (len(points) - 1)
        y = top_pad + chart_height * (1 - (value - min_value) / span)
        coords.append((x, y))

    path = " ".join(f"{'M' if i == 0 else 'L'} {x:.1f} {y:.1f}" for i, (x, y) in enumerate(coords))
    first_label = points[0][0][2:7]
    last_label = points[-1][0][2:7]
    return f"""
    <svg class="sparkline" viewBox="0 0 {width} {height}" role="img" aria-label="{escape(key)} trend">
      <line x1="{left_pad}" y1="{top_pad}" x2="{left_pad}" y2="{height - bottom_pad}" class="y-axis"></line>
      <line x1="{left_pad}" y1="{height - bottom_pad}" x2="{width - right_pad}" y2="{height - bottom_pad}" class="x-axis"></line>
      <text x="4" y="{top_pad + 4}" class="chart-label">{fmt_axis_value(max_value)}</text>
      <text x="4" y="{height - bottom_pad}" class="chart-label">{fmt_axis_value(min_value)}</text>
      <path d="{path}" fill="none" stroke="{color}" stroke-width="2.5"></path>
      <circle cx="{coords[-1][0]:.1f}" cy="{coords[-1][1]:.1f}" r="3.5" fill="{color}"></circle>
      <text x="{left_pad}" y="{height - 3}" class="chart-label">{first_label}</text>
      <text x="{width - right_pad}" y="{height - 3}" text-anchor="end" class="chart-label">{last_label}</text>
    </svg>
    """


def make_mini_bar_chart(rows, change_key, aria_label, bar_class="china-bar", width=360, height=118, limit=24):
    chart_rows = [row for row in rows if row.get(change_key) is not None][-limit:]
    changes = [row[change_key] for row in chart_rows]
    if not changes:
        return '<div class="empty-chart">暂无足够数据</div>'

    max_abs = max(abs(value) for value in changes) or 1
    left_pad = 42
    right_pad = 10
    top_pad = 14
    bottom_pad = 24
    chart_width = width - left_pad - right_pad
    chart_height = height - top_pad - bottom_pad
    zero_y = top_pad + chart_height / 2
    scale = (chart_height / 2 - 4) / max_abs
    slot_width = chart_width / len(chart_rows)
    bar_width = max(4, min(12, slot_width * 0.62))

    bars = []
    for i, row in enumerate(chart_rows):
        change = row[change_key]
        bar_height = max(1, abs(change) * scale)
        x = left_pad + slot_width * i + (slot_width - bar_width) / 2
        y = zero_y - bar_height if change >= 0 else zero_y
        bars.append(f"""
        <rect class="{bar_class}" x="{x:.1f}" y="{y:.1f}" width="{bar_width:.1f}" height="{bar_height:.1f}" rx="2"></rect>
        """)

    first_label = chart_rows[0]["date"][2:7]
    last_label = chart_rows[-1]["date"][2:7]
    return f"""
    <svg class="mini-bar-chart" viewBox="0 0 {width} {height}" role="img" aria-label="{escape(aria_label)}">
      <line x1="{left_pad}" y1="{top_pad}" x2="{left_pad}" y2="{height - bottom_pad}" class="y-axis"></line>
      <line x1="{left_pad}" y1="{zero_y:.1f}" x2="{width - right_pad}" y2="{zero_y:.1f}" class="x-axis"></line>
      <text x="4" y="{top_pad + 4}" class="chart-label">+{fmt_axis_value(max_abs)}</text>
      <text x="4" y="{height - bottom_pad}" class="chart-label">-{fmt_axis_value(max_abs)}</text>
      {''.join(bars)}
      <text x="{left_pad}" y="{height - 4}" class="chart-label">{first_label}</text>
      <text x="{width - right_pad}" y="{height - 4}" text-anchor="end" class="chart-label">{last_label}</text>
    </svg>
    """


def make_bar_chart(rows, change_key, aria_label, bar_class):
    chart_rows = [row for row in rows if row.get(change_key) is not None][-24:]
    changes = [row[change_key] for row in chart_rows]
    if not changes:
        return '<div class="empty-chart">暂无足够数据</div>'

    max_abs = max(abs(value) for value in changes) or 1
    bars = []
    x = 40
    width = 28
    gap = 10
    mid_y = 110
    scale = 78 / max_abs

    for row in chart_rows:
        change = row[change_key]
        height = abs(change) * scale
        y = mid_y - height if change >= 0 else mid_y
        label = row["date"][2:7]
        text_y = y - 6 if change >= 0 else y + height + 15
        bars.append(f"""
        <g>
          <rect class="{bar_class}" x="{x}" y="{y:.1f}" width="{width}" height="{height:.1f}" rx="3"></rect>
          <text x="{x + width / 2}" y="218" text-anchor="middle" class="chart-label">{label}</text>
          <text x="{x + width / 2}" y="{text_y:.1f}" text-anchor="middle" class="bar-value">{change:+.1f}</text>
        </g>
        """)
        x += width + gap

    svg_width = max(720, x + 20)
    return f"""
    <svg class="bar-chart" viewBox="0 0 {svg_width} 240" role="img" aria-label="{escape(aria_label)}">
      <line x1="30" y1="24" x2="30" y2="210" class="y-axis"></line>
      <line x1="30" y1="{mid_y}" x2="{svg_width - 20}" y2="{mid_y}" class="x-axis"></line>
      {''.join(bars)}
    </svg>
    """


def make_dual_axis_chart(
    rows,
    factor_key,
    factor_label,
    gold_key="gold_price",
    gold_label="黄金价格",
    width=720,
    height=220,
    limit=180,
    factor_color="#366b9f",
    gold_color="#b88728",
):
    points = [
        row for row in rows
        if row.get(factor_key) is not None and row.get(gold_key) is not None
    ][-limit:]
    if len(points) < 2:
        return '<div class="empty-chart">暂无足够双轴走势数据</div>'

    left_pad = 56
    right_pad = 66
    top_pad = 22
    bottom_pad = 32
    chart_width = width - left_pad - right_pad
    chart_height = height - top_pad - bottom_pad

    factor_values = [row[factor_key] for row in points]
    gold_values = [row[gold_key] for row in points]
    factor_min = min(factor_values)
    factor_max = max(factor_values)
    gold_min = min(gold_values)
    gold_max = max(gold_values)
    factor_span = factor_max - factor_min or 1
    gold_span = gold_max - gold_min or 1

    def x_for(index):
        return left_pad + chart_width * index / (len(points) - 1)

    def y_for(value, min_value, span):
        return top_pad + chart_height * (1 - (value - min_value) / span)

    factor_coords = [
        (x_for(i), y_for(row[factor_key], factor_min, factor_span))
        for i, row in enumerate(points)
    ]
    gold_coords = [
        (x_for(i), y_for(row[gold_key], gold_min, gold_span))
        for i, row in enumerate(points)
    ]
    factor_path = " ".join(
        f"{'M' if i == 0 else 'L'} {x:.1f} {y:.1f}" for i, (x, y) in enumerate(factor_coords)
    )
    gold_path = " ".join(
        f"{'M' if i == 0 else 'L'} {x:.1f} {y:.1f}" for i, (x, y) in enumerate(gold_coords)
    )
    first_label = points[0]["date"][2:7]
    last_label = points[-1]["date"][2:7]

    return f"""
    <div class="dual-axis-wrap">
      <div class="dual-axis-legend">
        <span><i style="background:{factor_color}"></i>{escape(factor_label)} {fmt_axis_value(points[-1][factor_key])}</span>
        <span><i style="background:{gold_color}"></i>{escape(gold_label)} {fmt_axis_value(points[-1][gold_key])}</span>
      </div>
      <svg class="dual-axis-chart" viewBox="0 0 {width} {height}" role="img" aria-label="{escape(factor_label)}与{escape(gold_label)}双轴走势">
        <line x1="{left_pad}" y1="{top_pad}" x2="{left_pad}" y2="{height - bottom_pad}" class="left-axis"></line>
        <line x1="{width - right_pad}" y1="{top_pad}" x2="{width - right_pad}" y2="{height - bottom_pad}" class="right-axis"></line>
        <line x1="{left_pad}" y1="{height - bottom_pad}" x2="{width - right_pad}" y2="{height - bottom_pad}" class="x-axis"></line>
        <text x="6" y="{top_pad + 4}" class="chart-label">{fmt_axis_value(factor_max)}</text>
        <text x="6" y="{height - bottom_pad}" class="chart-label">{fmt_axis_value(factor_min)}</text>
        <text x="{width - 6}" y="{top_pad + 4}" text-anchor="end" class="chart-label">{fmt_axis_value(gold_max)}</text>
        <text x="{width - 6}" y="{height - bottom_pad}" text-anchor="end" class="chart-label">{fmt_axis_value(gold_min)}</text>
        <path d="{factor_path}" fill="none" stroke="{factor_color}" stroke-width="2.3"></path>
        <path d="{gold_path}" fill="none" stroke="{gold_color}" stroke-width="2.3"></path>
        <circle cx="{factor_coords[-1][0]:.1f}" cy="{factor_coords[-1][1]:.1f}" r="3.1" fill="{factor_color}"></circle>
        <circle cx="{gold_coords[-1][0]:.1f}" cy="{gold_coords[-1][1]:.1f}" r="3.1" fill="{gold_color}"></circle>
        <text x="{left_pad}" y="{height - 8}" class="chart-label">{first_label}</text>
        <text x="{width - right_pad}" y="{height - 8}" text-anchor="end" class="chart-label">{last_label}</text>
      </svg>
    </div>
    """


def make_corr_chart(series_defs, aria_label, width=720, height=180, limit=96):
    top_pad = 18
    bottom_pad = 34
    left_pad = 38
    right_pad = 16
    chart_width = width - left_pad - right_pad
    chart_height = height - top_pad - bottom_pad

    def y_for(value):
        clipped = max(-1, min(1, value))
        return top_pad + chart_height * (1 - (clipped + 1) / 2)

    zero_y = y_for(0)
    paths = []
    legends = []
    visible_points = []
    for series in series_defs:
        points = [row for row in series["rows"] if row.get("corr") is not None][-limit:]
        if len(points) < 2:
            continue
        visible_points.extend(points)
        coords = []
        for i, row in enumerate(points):
            x = left_pad + chart_width * i / (len(points) - 1)
            coords.append((x, y_for(row["corr"])))
        path = " ".join(f"{'M' if i == 0 else 'L'} {x:.1f} {y:.1f}" for i, (x, y) in enumerate(coords))
        dash = ' stroke-dasharray="5 4"' if series.get("dash") else ""
        paths.append(
            f'<path d="{path}" fill="none" stroke="{series["color"]}" stroke-width="2.4"{dash}></path>'
            f'<circle cx="{coords[-1][0]:.1f}" cy="{coords[-1][1]:.1f}" r="3.2" fill="{series["color"]}"></circle>'
        )
        legends.append(
            f'<span><i style="background:{series["color"]}"></i>{escape(series["label"])} {fmt_corr(points[-1]["corr"])}</span>'
        )

    if not paths:
        return '<div class="empty-chart">暂无足够滚动相关数据</div>'

    visible_points.sort(key=lambda row: row["_date"])
    first_label = visible_points[0]["date"][2:7]
    last_label = visible_points[-1]["date"][2:7]

    return f"""
    <div class="corr-wrap">
      <div class="corr-legend">{''.join(legends)}</div>
      <svg class="corr-chart" viewBox="0 0 {width} {height}" role="img" aria-label="{escape(aria_label)}">
        <line x1="{left_pad}" y1="{zero_y:.1f}" x2="{width - right_pad}" y2="{zero_y:.1f}" class="zero-axis"></line>
        <line x1="{left_pad}" y1="{top_pad}" x2="{left_pad}" y2="{height - bottom_pad}" class="y-axis"></line>
        <line x1="{left_pad}" y1="{height - bottom_pad}" x2="{width - right_pad}" y2="{height - bottom_pad}" class="x-axis"></line>
        <text x="8" y="{top_pad + 4}" class="chart-label">+1</text>
        <text x="8" y="{zero_y + 4:.1f}" class="chart-label">0</text>
        <text x="8" y="{height - bottom_pad}" class="chart-label">-1</text>
        <text x="{left_pad}" y="{height - 8}" class="chart-label">{first_label}</text>
        <text x="{width - right_pad}" y="{height - 8}" text-anchor="end" class="chart-label">{last_label}</text>
        {''.join(paths)}
      </svg>
    </div>
    """


def tone_badge(tone):
    if tone == "有效":
        klass = "supportive"
    elif tone == "失效":
        klass = "headwind"
    else:
        klass = "neutral"
    return f'<span class="tone tone-{klass}">{escape(tone)}</span>'


def make_central_bank_relationship_card(relationship):
    weak_periods = relationship.get("weak_periods") or []
    if weak_periods:
        weak_text = "；".join(f"{start} 至 {end}" for start, end in weak_periods)
    else:
        weak_text = "未识别到连续 3 个月以上的低相关段"

    phase_rows = "".join(
        f"""
        <tr>
          <td>{escape(phase['label'])}</td>
          <td>{fmt_corr(phase['corr'])}</td>
          <td>{tone_badge(phase['tone'])}</td>
          <td>{fmt_pct(phase['gold_return'])}</td>
          <td>{phase['observations']}</td>
        </tr>
        """
        for phase in relationship["phases"]
    )
    chart = make_corr_chart(
        [{"label": "24个月滚动相关", "rows": relationship["rolling_corr"], "color": "#237a57"}],
        "央行购金与黄金收益滚动相关",
        limit=84,
    )
    trend_chart = make_dual_axis_chart(
        relationship["trend_rows"],
        relationship["factor_key"],
        relationship["factor_label"],
        relationship["gold_key"],
        limit=84,
        factor_color="#237a57",
    )

    return f"""
    <article class="relationship-card">
      <div class="relationship-head">
        <div>
          <div class="eyebrow">关系检验</div>
          <h3>{escape(relationship['name'])}</h3>
        </div>
        {tone_badge(relationship['latest_tone'])}
      </div>
      <div class="metric-strip">
        <div><span>起始显著推动</span><strong>{escape(relationship.get('start_month') or '—')}</strong></div>
        <div><span>当前滚动相关</span><strong>{fmt_corr(relationship['latest_corr'])}</strong></div>
        <div><span>最近 3 个月</span><strong>{fmt_signed(relationship['latest']['factor_change'])}</strong></div>
      </div>
      <p class="relationship-read">{escape(relationship['read'])} 什么时候不好用：{escape(weak_text)}。</p>
      <h4>双轴走势</h4>
      {trend_chart}
      <h4>滚动相关</h4>
      {chart}
      <h4>阶段表现</h4>
      <table class="phase-table">
        <thead><tr><th>阶段</th><th>相关系数</th><th>判断</th><th>黄金收益</th><th>样本(月)</th></tr></thead>
        <tbody>{phase_rows}</tbody>
      </table>
    </article>
    """


def make_horizon_relationship_card(relationship):
    weak_periods = relationship.get("weak_periods") or []
    if weak_periods:
        weak_text = "；".join(f"{start} 至 {end}" for start, end in weak_periods[-2:])
    else:
        weak_text = "未识别到连续低相关段"

    phase_rows = "".join(
        f"""
        <tr>
          <td>{escape(phase['label'])}</td>
          <td>{fmt_corr(phase['short_corr'])} {tone_badge(phase['short_tone'])}</td>
          <td>{fmt_corr(phase['medium_corr'])} {tone_badge(phase['medium_tone'])}</td>
          <td>{fmt_pct(phase['gold_return'])}</td>
          <td>{phase['observations']}</td>
        </tr>
        """
        for phase in relationship["phases"]
    )
    chart = make_corr_chart(
        [
            {
                "label": relationship["short_term"]["label"],
                "rows": relationship["short_term"]["rolling_corr"],
                "color": "#366b9f",
            },
            {
                "label": relationship["medium_term"]["label"],
                "rows": relationship["medium_term"]["rolling_corr"],
                "color": "#b88728",
                "dash": True,
            },
        ],
        f"{relationship['name']}与黄金收益滚动相关",
        limit=180,
    )
    trend_chart = make_dual_axis_chart(
        relationship["trend_rows"],
        relationship["factor_key"],
        relationship["factor_label"],
        relationship["gold_key"],
        limit=180,
    )

    return f"""
    <article class="relationship-card">
      <div class="relationship-head">
        <div>
          <div class="eyebrow">关系检验</div>
          <h3>{escape(relationship['name'])}</h3>
        </div>
        {tone_badge(relationship['latest_tone'])}
      </div>
      <div class="metric-strip">
        <div><span>{escape(relationship['short_term']['label'])}</span><strong>{fmt_corr(relationship['short_term']['latest_corr'])}</strong></div>
        <div><span>{escape(relationship['medium_term']['label'])}</span><strong>{fmt_corr(relationship['medium_term']['latest_corr'])}</strong></div>
        <div><span>低相关窗口</span><strong>{escape(weak_text)}</strong></div>
      </div>
      <p class="relationship-read">{escape(relationship['read'])}</p>
      <h4>双轴走势</h4>
      {trend_chart}
      <h4>滚动相关</h4>
      {chart}
      <h4>阶段表现</h4>
      <table class="phase-table">
        <thead><tr><th>阶段</th><th>短期相关</th><th>中期相关</th><th>黄金收益</th><th>样本(日)</th></tr></thead>
        <tbody>{phase_rows}</tbody>
      </table>
    </article>
    """


def make_positioning_relationship_card(relationship):
    phase_rows = []
    for sub_metric in relationship["sub_metrics"]:
        for phase in sub_metric["phases"]:
            phase_rows.append(f"""
        <tr>
          <td>{escape(sub_metric['name'])}</td>
          <td>{escape(phase['label'])}</td>
          <td>{fmt_corr(phase['corr'])}</td>
          <td>{tone_badge(phase['tone'])}</td>
          <td>{fmt_pct(phase['gold_return'])}</td>
          <td>{phase['observations']}</td>
        </tr>
            """)

    metric_strip = "".join(
        f"""
        <div><span>{escape(sub_metric['name'])}</span><strong>{fmt_corr(sub_metric['latest_corr'])}</strong></div>
        """
        for sub_metric in relationship["sub_metrics"]
    )
    chart = make_corr_chart(
        [
            {
                "label": relationship["sub_metrics"][0]["name"],
                "rows": relationship["sub_metrics"][0]["rolling_corr"],
                "color": "#237a57",
            },
            {
                "label": relationship["sub_metrics"][1]["name"],
                "rows": relationship["sub_metrics"][1]["rolling_corr"],
                "color": "#366b9f",
                "dash": True,
            },
            {
                "label": relationship["sub_metrics"][2]["name"],
                "rows": relationship["sub_metrics"][2]["rolling_corr"],
                "color": "#b88728",
            },
        ],
        "仓位与技术指标与黄金收益滚动相关",
        limit=120,
    )
    trend_chart_items = []
    for sub_metric in relationship["sub_metrics"]:
        trend_chart = make_dual_axis_chart(
            sub_metric["trend_rows"],
            sub_metric["factor_key"],
            sub_metric["factor_label"],
            sub_metric["gold_key"],
            limit=120,
        )
        trend_chart_items.append(f"""
        <div class="dual-axis-item">
          <h5>{escape(sub_metric['name'])}</h5>
          {trend_chart}
        </div>
        """)
    trend_charts = "".join(trend_chart_items)

    return f"""
    <article class="relationship-card relationship-card-wide">
      <div class="relationship-head">
        <div>
          <div class="eyebrow">关系检验</div>
          <h3>{escape(relationship['name'])}</h3>
        </div>
        <span class="tone tone-neutral">{escape(relationship['latest_tone'])}</span>
      </div>
      <div class="metric-strip">{metric_strip}</div>
      <p class="relationship-read">{escape(relationship['read'])}</p>
      <h4>双轴走势</h4>
      <div class="dual-axis-grid">{trend_charts}</div>
      <h4>滚动相关</h4>
      {chart}
      <h4>阶段表现</h4>
      <table class="phase-table">
        <thead><tr><th>子指标</th><th>阶段</th><th>相关系数</th><th>判断</th><th>黄金收益</th><th>样本</th></tr></thead>
        <tbody>{''.join(phase_rows)}</tbody>
      </table>
    </article>
    """


def make_relationship_section(relationships):
    cards = []
    for relationship in relationships:
        if relationship["id"] in {"central_bank_purchases", "epu", "gpr"}:
            cards.append(make_central_bank_relationship_card(relationship))
        elif relationship["id"] in {"real_rate", "dollar", "inflation_expectation", "price_trend"}:
            cards.append(make_horizon_relationship_card(relationship))
        elif relationship["id"] == "positioning_technical":
            cards.append(make_positioning_relationship_card(relationship))

    return f"""
  <section class="wide relationship-section">
    <div class="section-title-row">
      <div>
        <h2>因子关系检验</h2>
        <p>用同一批数值数据检验“指标是否好用”：看当前滚动相关，也看阶段表现。正负方向按常识方向设定，但结论只来自相关系数。</p>
      </div>
    </div>
    <div class="relationship-grid">
      {''.join(cards)}
    </div>
  </section>
    """


def layer_card(layer):
    chart = ""
    if layer["chart_key"]:
        if layer.get("chart_type") == "bar":
            chart = make_mini_bar_chart(
                layer["chart_rows"],
                layer["chart_key"],
                layer.get("chart_label", f"{layer['name']}环比柱状图"),
            )
        else:
            chart = make_sparkline(layer["chart_rows"], layer["chart_key"])

    return f"""
    <article class="layer-card state-{layer['state']}">
      <div class="layer-head">
        <div>
          <div class="eyebrow">{escape(layer['name'])}</div>
          <h3>{escape(layer['value_label'])}</h3>
        </div>
        <div class="badges">
          <span class="state">{state_badge(layer['state'])}</span>
          <span class="quality">{quality_badge(layer['data_quality'])}</span>
        </div>
      </div>
      <p>{escape(layer['read'])}</p>
      {chart}
      <dl class="mini-meta">
        <div><dt>数据日期</dt><dd>{escape(layer['latest']['date'] or '—')}</dd></div>
        <div><dt>最近变化</dt><dd>{escape(layer['change_label'])}</dd></div>
        <div><dt>来源</dt><dd>{escape(layer['source'])}</dd></div>
      </dl>
    </article>
    """


def make_change_list(layers):
    items = []
    for layer in layers:
        if layer["state"] == "planned":
            items.append(f"{layer['name']}：{layer['read']}")
        else:
            items.append(f"{layer['name']}：{layer['read']}")
    return items[:5]


def make_price_summary_section(technical_layer):
    latest = technical_layer["latest"]
    technical = technical_layer["technical"]
    chart = make_gold_chart_panel(technical_layer)
    return f"""
    <section id="gold-price-section" class="price-stage">
      <div class="price-panel">
        <div class="section-eyebrow">黄金现货 · 美元/盎司</div>
        <div class="price-value">{fmt_number(latest['gold_price'])}</div>
        <div class="price-move">近 5 个有效观测日 {fmt_pct(latest.get('return_5d'))}</div>
        {chart}
        <p class="chart-caption">数据日期 {escape(latest.get('date') or '—')} · 来源 {escape(technical_layer['source'])}</p>
      </div>
      <aside class="technical-panel state-{escape(technical_layer['state'])}">
        <div class="section-eyebrow">技术状态</div>
        <h2>{escape(technical['medium_term'])} · {escape(technical['short_term'])}</h2>
        <p>{escape(technical['alignment'])}</p>
        <dl class="ma-summary">
          <div><dt>MA20</dt><dd>{fmt_number(latest.get('ma20'))} · {fmt_pct(latest.get('gap_ma20'))}</dd></div>
          <div><dt>MA60</dt><dd>{fmt_number(latest.get('ma60'))} · {fmt_pct(latest.get('gap_ma60'))}</dd></div>
          <div><dt>MA200</dt><dd>{fmt_number(latest.get('ma200'))} · {fmt_pct(latest.get('gap_ma200'))}</dd></div>
        </dl>
        <p class="technical-trigger"><strong>判断改变条件：</strong>{escape(technical['trigger'])}</p>
      </aside>
    </section>
    """


def make_driver_section(rows):
    body = "".join(
        f"""
      <div class="driver-row" data-driver-id="{escape(row['id'])}">
        <span class="driver-state state-{escape(row['state'])}">{escape(state_badge(row['state']))}</span>
        <div><strong>{escape(row['name'])}</strong><small>{escape(row['category'])}</small></div>
        <strong>{escape(row['value'])}</strong>
        <span>{escape(row['change'])}</span>
        <span>{escape(row['read'])}</span>
        <span class="quality">{escape(quality_badge(row['quality']))}</span>
      </div>
        """
        for row in rows
    )
    return f"""
    <section class="driver-panel">
      <div class="section-heading"><h2>当前主要驱动</h2><span>按决策相关性排序</span></div>
      <div class="driver-head" aria-hidden="true">
        <span>状态</span><span>驱动</span><span>当前值</span><span>近期变化</span><span>解读</span><span>质量</span>
      </div>
      <div id="driver-table" class="driver-table">{body}</div>
    </section>
    """


def make_recent_changes_section(changes):
    cards = "".join(
        f"""
      <article class="recent-change tone-{escape(item['tone'])}">
        <div class="change-meta">
          <small>{escape(item['label'])}</small>
          <span class="change-state state-{escape(item['tone'])}">{escape(state_badge(item['tone']))}</span>
        </div>
        <h3>{escape(item['headline'])}</h3>
        <p>{escape(item['detail'])}</p>
      </article>
        """
        for item in changes
    )
    return f"""
    <aside id="recent-changes" class="recent-panel">
      <h2>最近发生了什么</h2>
      {cards}
    </aside>
    """


def make_research_details(dashboard):
    relationship_section = make_relationship_section(dashboard["relationships"])
    china_chart = make_bar_chart(
        dashboard["reserve_rows"],
        "china_mom_change",
        "中国央行黄金储备环比变化柱状图",
        "china-bar",
    )
    global_chart = make_bar_chart(
        dashboard["reserve_rows"],
        "global_mom_change",
        "全球央行黄金储备环比变化柱状图",
        "global-bar",
    )
    reserve_layer = next(
        layer for layer in dashboard["layers"] if layer["id"] == "official_reserves")
    positioning_layer = next(
        layer
        for layer in dashboard["layers"]
        if layer["id"] == "positioning_technical"
    )
    reserve_sources = (
        f"中国：{reserve_layer['latest'].get('china_source') or '—'}；"
        f"全球：{reserve_layer['latest'].get('global_source') or '—'}"
    )
    quality_rows = "".join(
        f"""
      <tr>
        <td>{escape(layer['name'])}</td>
        <td>{escape(quality_badge(layer['data_quality']))}</td>
        <td>{'—' if layer.get('lag_days') is None else str(layer['lag_days']) + ' 天'}</td>
        <td>{escape(layer['latest'].get('date') or '—')}</td>
        <td>{escape(layer['source'])}</td>
      </tr>
        """
        for layer in dashboard["layers"]
    )
    valuation = dashboard["valuation"] or {}
    valuation_text = (
        f"长期估值：金价/M2 {fmt_number(valuation.get('gold_to_m2'), 3)}，"
        f"估值分位 {fmt_number((valuation.get('valuation_percentile') or 0) * 100)}%，"
        f"数据日期 {valuation.get('date', '—')}。"
    )
    scoring_summary = (
        f"驱动净分 {int(dashboard['score']):+d} / "
        f"有效驱动 {dashboard['active_layers']} 组 · "
        f"归一化倾向 {dashboard['tendency']:+.2f} · "
        f"当前姿态 {dashboard['posture']}"
    )
    return f"""
    <details class="research-evidence">
      <summary><strong>研究依据</strong><span>关系检验 · 历史阶段 · 数据来源与方法</span></summary>
      {relationship_section}
      <section class="wide">
        <h2>央行购金历史</h2>
        <p class="research-intro">{escape(reserve_layer['read'])}</p>
        <div class="chart-pair">
          <div class="chart-box"><h3>中国央行黄金储备环比（吨）</h3>{china_chart}</div>
          <div class="chart-box"><h3>全球官方黄金储备环比（吨）</h3>{global_chart}</div>
        </div>
        <p class="method">数据日期 {escape(reserve_layer['latest'].get('date') or '—')} · 来源 {escape(reserve_sources)}</p>
      </section>
      <section class="wide">
        <h2>数据质量与来源</h2>
        <table>
          <thead><tr><th>层级</th><th>质量</th><th>滞后</th><th>数据日期</th><th>来源</th></tr></thead>
          <tbody>{quality_rows}</tbody>
        </table>
        <p class="method">仓位数据说明：{escape(positioning_layer['read'])}</p>
        <p class="method">{escape(valuation_text)}</p>
        <p class="method">生成时间：{escape(dashboard['updated_at'])} · 数据文件：{escape(dashboard['source_file'])}</p>
      </section>
      <section class="wide">
        <h2>计算方法与已知局限</h2>
        <p class="method"><strong>当前计分：</strong>{escape(scoring_summary)}</p>
        <ul>
          <li>五组驱动为实际利率、美元、中国央行购金、通胀预期、资金与仓位；ETF 与 CFTC 合为一组，价格技术不参与驱动合成。</li>
          <li>支持记 +1、中性记 0、压力记 -1；归一化倾向 +0.25 及以上为偏多，-0.25 及以下为承压，其余为中性。缺失和严重滞后不计分，滞后仍计分；有效驱动少于 3 组时显示数据不足。</li>
          <li>技术状态独立判断；只有价格与 MA20、MA60、MA200 满足完整次序时，才称为严格多头或空头排列。</li>
          <li>该姿态是等权启发式框架，实际利率、美元与通胀预期可能共线；关系检验只说明统计相关，相关性不代表因果，滚动相关使用重叠窗口。</li>
          <li>EPU 与 GPR 只保留在研究依据，不参与首页姿态；SAFE 官方序列仍需手工维护。</li>
        </ul>
      </section>
    </details>
    """


def build_html(dashboard):
    scored_states = [
        layer["state"]
        for layer in dashboard["driver_layers"]
        if layer["state"] in {"supportive", "neutral", "headwind"}
        and layer.get("data_quality") != "very-stale"
    ]
    summary = (
        f"{scored_states.count('headwind')} 项压力 · "
        f"{scored_states.count('supportive')} 项支持 · "
        f"{scored_states.count('neutral')} 项中性"
    )
    updated = f"主要市场数据截至 {dashboard['technical_layer']['latest']['date']}"
    total_drivers = len(dashboard["driver_layers"])
    excluded_drivers = total_drivers - dashboard["active_layers"]
    availability = (
        f"{dashboard['active_layers']} / {total_drivers} 组驱动可用 · "
        f"{excluded_drivers} 组过期或缺失"
    )
    price_section = make_price_summary_section(dashboard["technical_layer"])
    driver_section = make_driver_section(dashboard["driver_rows"])
    recent_changes_section = make_recent_changes_section(dashboard["recent_changes"])
    research_details = make_research_details(dashboard)

    html = f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{escape(dashboard['title'])}</title>
  <style>
    :root {{
      color-scheme: light;
      --bg: #f7f8f5;
      --panel: #ffffff;
      --line: #dfe3da;
      --text: #172019;
      --muted: #677064;
      --support: #237a57;
      --neutral: #7a6a26;
      --headwind: #b94343;
      --planned: #687080;
      --gold: #b88728;
      --blue: #366b9f;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      background: var(--bg);
      color: var(--text);
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
      line-height: 1.55;
    }}
    main {{
      max-width: 1180px;
      margin: 0 auto;
      padding: 28px 20px 48px;
    }}
    h1, h2, h3, p {{ margin: 0; }}
    .hero {{
      border-bottom: 1px solid var(--line);
      padding-bottom: 20px;
      display: grid;
      grid-template-columns: minmax(0, 1fr) 220px;
      gap: 20px;
      align-items: end;
    }}
    .kicker, .eyebrow {{
      color: var(--muted);
      font-size: 12px;
      font-weight: 700;
      letter-spacing: .08em;
      text-transform: uppercase;
    }}
    h1 {{
      font-size: 34px;
      line-height: 1.15;
      margin-top: 6px;
      letter-spacing: 0;
    }}
    .subtitle {{
      margin-top: 8px;
      color: var(--muted);
      max-width: 760px;
      font-size: 14px;
    }}
    .posture {{
      border: 1px solid var(--line);
      border-radius: 8px;
      background: var(--panel);
      padding: 14px;
      text-align: right;
    }}
    .posture .label {{ color: var(--muted); font-size: 12px; }}
    .posture .value {{ font-size: 30px; font-weight: 800; }}
    .posture .score {{ color: var(--muted); font-family: ui-monospace, SFMono-Regular, Menlo, monospace; }}
    .cards {{
      display: grid;
      grid-template-columns: repeat(5, minmax(0, 1fr));
      gap: 12px;
      margin-top: 20px;
    }}
    .layer-card {{
      min-width: 0;
      border: 1px solid var(--line);
      border-top: 3px solid var(--planned);
      border-radius: 8px;
      background: var(--panel);
      padding: 14px;
      display: flex;
      flex-direction: column;
      gap: 10px;
    }}
    .state-supportive {{ border-top-color: var(--support); }}
    .state-neutral {{ border-top-color: var(--neutral); }}
    .state-headwind {{ border-top-color: var(--headwind); }}
    .state-planned {{ border-top-color: var(--planned); }}
    .layer-head {{
      display: flex;
      justify-content: space-between;
      gap: 10px;
      align-items: start;
    }}
    .layer-card h3 {{
      font-size: 18px;
      line-height: 1.2;
      margin-top: 4px;
    }}
    .layer-card p {{
      color: #334036;
      font-size: 13px;
      min-height: 60px;
    }}
    .badges {{
      display: flex;
      flex-direction: column;
      gap: 4px;
      align-items: flex-end;
      flex-shrink: 0;
    }}
    .state, .quality {{
      border: 1px solid var(--line);
      border-radius: 6px;
      padding: 2px 7px;
      font-size: 11px;
      color: var(--muted);
      background: #f7f8f5;
      white-space: nowrap;
    }}
    .mini-meta {{
      display: grid;
      gap: 6px;
      margin: 0;
      padding-top: 8px;
      border-top: 1px solid var(--line);
      font-size: 11px;
    }}
    .mini-meta div {{
      display: grid;
      grid-template-columns: 58px 1fr;
      gap: 8px;
    }}
    dt {{ color: var(--muted); }}
    dd {{ margin: 0; min-width: 0; overflow-wrap: anywhere; }}
    .section-grid {{
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      gap: 14px;
      margin-top: 18px;
    }}
    .panel {{
      border: 1px solid var(--line);
      border-radius: 8px;
      background: var(--panel);
      padding: 16px;
    }}
    .panel h2 {{
      font-size: 16px;
      margin-bottom: 10px;
    }}
    ul {{
      margin: 0;
      padding-left: 18px;
      color: #334036;
      font-size: 13px;
    }}
    li + li {{ margin-top: 6px; }}
    .wide {{
      margin-top: 18px;
      border: 1px solid var(--line);
      border-radius: 8px;
      background: var(--panel);
      padding: 16px;
      overflow-x: auto;
    }}
    .wide h2 {{ font-size: 16px; margin-bottom: 10px; }}
    .chart-pair {{
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 14px;
    }}
    .chart-box {{
      min-width: 520px;
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 12px;
    }}
    .chart-box h3 {{
      font-size: 14px;
      margin-bottom: 6px;
    }}
    .sparkline, .mini-bar-chart, .bar-chart, .corr-chart, .dual-axis-chart {{
      width: 100%;
      height: auto;
      display: block;
    }}
    .section-title-row {{
      display: flex;
      justify-content: space-between;
      gap: 16px;
      margin-bottom: 14px;
    }}
    .section-title-row p {{
      color: var(--muted);
      font-size: 13px;
      margin-top: 4px;
      max-width: 760px;
    }}
    .relationship-grid {{
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 14px;
    }}
    .relationship-card {{
      min-width: 0;
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 14px;
      background: #fcfdfb;
    }}
    .relationship-card-wide {{
      grid-column: 1 / -1;
    }}
    .relationship-head {{
      display: flex;
      align-items: flex-start;
      justify-content: space-between;
      gap: 12px;
      margin-bottom: 12px;
    }}
    .relationship-head h3 {{
      font-size: 18px;
      margin-top: 3px;
    }}
    .relationship-card h4 {{
      margin: 14px 0 6px;
      font-size: 13px;
      color: var(--muted);
    }}
    .relationship-read {{
      color: #334036;
      font-size: 13px;
      margin-top: 10px;
    }}
    .metric-strip {{
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      gap: 8px;
    }}
    .metric-strip div {{
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 9px;
      min-width: 0;
      background: #ffffff;
    }}
    .metric-strip span {{
      display: block;
      color: var(--muted);
      font-size: 11px;
      margin-bottom: 3px;
    }}
    .metric-strip strong {{
      display: block;
      font-size: 16px;
      line-height: 1.2;
      overflow-wrap: anywhere;
    }}
    .tone {{
      display: inline-block;
      border-radius: 6px;
      padding: 2px 7px;
      font-size: 11px;
      border: 1px solid var(--line);
      white-space: nowrap;
    }}
    .tone-supportive {{ color: var(--support); background: #ecf6f0; }}
    .tone-neutral {{ color: var(--neutral); background: #faf6e6; }}
    .tone-headwind {{ color: var(--headwind); background: #fbeeee; }}
    .corr-wrap {{
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 10px;
      background: #ffffff;
    }}
    .corr-legend {{
      display: flex;
      gap: 14px;
      flex-wrap: wrap;
      color: var(--muted);
      font-size: 12px;
      margin-bottom: 4px;
    }}
    .corr-legend span {{
      display: inline-flex;
      align-items: center;
      gap: 5px;
    }}
    .corr-legend i {{
      display: inline-block;
      width: 16px;
      height: 3px;
      border-radius: 4px;
    }}
    .dual-axis-wrap {{
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 10px;
      background: #ffffff;
    }}
    .dual-axis-grid {{
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      gap: 10px;
    }}
    .dual-axis-item {{
      min-width: 0;
    }}
    .dual-axis-item h5 {{
      margin: 0 0 6px;
      color: var(--muted);
      font-size: 12px;
    }}
    .dual-axis-legend {{
      display: flex;
      gap: 14px;
      flex-wrap: wrap;
      color: var(--muted);
      font-size: 12px;
      margin-bottom: 4px;
    }}
    .dual-axis-legend span {{
      display: inline-flex;
      align-items: center;
      gap: 5px;
    }}
    .dual-axis-legend i {{
      display: inline-block;
      width: 16px;
      height: 3px;
      border-radius: 4px;
    }}
    .phase-table td:nth-child(2),
    .phase-table td:nth-child(3) {{
      white-space: nowrap;
    }}
    .x-axis, .y-axis, .zero-axis, .left-axis, .right-axis {{ stroke: #c5cbbf; stroke-width: 1; }}
    .right-axis {{ stroke: var(--gold); }}
    .left-axis {{ stroke: var(--blue); }}
    .zero-axis {{ stroke-dasharray: 3 4; }}
    .chart-label, .bar-value {{
      fill: var(--muted);
      font-size: 11px;
    }}
    .china-bar {{ fill: var(--support); }}
    .global-bar {{ fill: var(--gold); }}
    table {{
      border-collapse: collapse;
      width: 100%;
      font-size: 13px;
    }}
    th, td {{
      border-bottom: 1px solid var(--line);
      padding: 9px 8px;
      text-align: left;
      vertical-align: top;
    }}
    th {{
      color: var(--muted);
      font-size: 12px;
      font-weight: 700;
    }}
    .method {{
      margin-top: 18px;
      color: var(--muted);
      font-size: 12px;
    }}
    .empty-chart {{
      min-height: 96px;
      color: var(--muted);
      display: flex;
      align-items: center;
      justify-content: center;
      border: 1px dashed var(--line);
      border-radius: 8px;
      font-size: 12px;
    }}
    @media (max-width: 980px) {{
      .hero, .section-grid, .chart-pair {{ grid-template-columns: 1fr; }}
      .relationship-grid {{ grid-template-columns: 1fr; }}
      .dual-axis-grid {{ grid-template-columns: 1fr; }}
      .cards {{ grid-template-columns: repeat(2, minmax(0, 1fr)); }}
      .posture {{ text-align: left; }}
      .chart-box {{ min-width: 0; }}
    }}
    @media (max-width: 560px) {{
      main {{ padding: 20px 14px 36px; }}
      h1 {{ font-size: 28px; }}
      .cards {{ grid-template-columns: 1fr; }}
    }}
    .decision-header {{
      display: grid;
      grid-template-columns: minmax(0, 1fr) 220px;
      gap: 20px;
      align-items: end;
      border-bottom: 1px solid var(--line);
      padding-bottom: 18px;
    }}
    .decision-header .kicker {{
      font-size: 13px;
      text-transform: none;
    }}
    .decision-header p {{
      margin-top: 8px;
      color: var(--muted);
      font-size: 14px;
    }}
    .posture {{
      border-radius: 14px;
      padding: 16px;
      border-left: 4px solid var(--neutral);
    }}
    .posture.state-supportive {{ border-left-color: var(--support); }}
    .posture.state-headwind {{ border-left-color: var(--headwind); }}
    .posture span, .posture strong, .posture small {{ display: block; }}
    .posture span {{ color: var(--muted); font-size: 13px; }}
    .posture strong {{ font-size: 28px; line-height: 1.2; margin: 3px 0 5px; }}
    .posture small {{ color: var(--muted); font-size: 12px; }}
    .price-stage {{
      display: grid;
      grid-template-columns: minmax(0, 1.8fr) minmax(270px, .8fr);
      gap: 16px;
      margin-top: 18px;
    }}
    .price-panel, .technical-panel, .driver-panel, .recent-panel, .research-evidence {{
      border: 1px solid var(--line);
      border-radius: 14px;
      background: var(--panel);
    }}
    .price-panel, .technical-panel, .driver-panel, .recent-panel {{ padding: 18px; }}
    .section-eyebrow {{
      color: var(--muted);
      font-size: 12px;
      font-weight: 700;
      letter-spacing: .06em;
    }}
    .price-value {{ font-size: 36px; font-weight: 800; letter-spacing: -.03em; }}
    .price-move {{ color: var(--muted); margin: 2px 0 10px; font-size: 14px; }}
    .chart-caption {{ color: var(--muted); margin-top: 9px; font-size: 12px; }}
    .gold-chart-panel {{ margin-top: 12px; }}
    .chart-heading, .chart-controls, .range-controls, .series-controls, .chart-tooltip {{
      display: flex;
      align-items: center;
      gap: 8px;
      flex-wrap: wrap;
    }}
    .chart-heading {{ justify-content: space-between; margin-bottom: 8px; }}
    .chart-heading strong {{ font-size: 14px; }}
    .chart-heading span {{ color: var(--muted); font-size: 12px; }}
    .chart-controls {{ justify-content: space-between; gap: 10px; }}
    .chart-controls button {{
      border: 1px solid var(--line);
      border-radius: 999px;
      background: #f7f8f5;
      color: var(--text);
      padding: 5px 10px;
      font: inherit;
      font-size: 12px;
      cursor: pointer;
    }}
    .chart-controls button[aria-pressed="true"] {{
      background: #2f5e4c;
      color: #fff;
      border-color: #2f5e4c;
    }}
    .chart-controls button:focus-visible {{ outline: 2px solid var(--blue); outline-offset: 2px; }}
    .chart-controls button:disabled {{ cursor: default; opacity: .45; }}
    .chart-stack {{ margin-top: 8px; }}
    .gold-chart {{ width: 100%; height: auto; display: block; touch-action: pan-y; }}
    .gold-chart[hidden] {{ display: none; }}
    .chart-series {{ fill: none; stroke-width: 2; vector-effect: non-scaling-stroke; }}
    .series-gold_price {{ stroke: var(--gold); stroke-width: 3; }}
    .series-ma20 {{ stroke: var(--support); stroke-dasharray: 5 4; }}
    .series-ma60 {{ stroke: var(--headwind); stroke-dasharray: 3 4; }}
    .series-ma200 {{ stroke: var(--blue); stroke-dasharray: 7 4; }}
    .hide-ma20 .series-ma20,
    .hide-ma60 .series-ma60,
    .hide-ma200 .series-ma200 {{ display: none; }}
    .hover-line {{ stroke: var(--muted); stroke-dasharray: 2 3; }}
    .hover-price {{ fill: var(--gold); stroke: #fff; stroke-width: 2; }}
    .hover-line[hidden], .hover-price[hidden] {{ display: none; }}
    .chart-empty {{ fill: var(--muted); font-size: 13px; }}
    .chart-tooltip {{ color: var(--muted); font-size: 12px; min-height: 20px; }}
    .chart-tooltip strong, .chart-tooltip b {{ color: var(--text); }}
    .technical-panel {{ border-left: 4px solid var(--neutral); }}
    .technical-panel.state-supportive {{ border-left-color: var(--support); }}
    .technical-panel.state-headwind {{ border-left-color: var(--headwind); }}
    .technical-panel h2 {{ margin-top: 6px; font-size: 21px; line-height: 1.3; }}
    .technical-panel > p {{ margin-top: 7px; color: #334036; }}
    .ma-summary {{ display: grid; gap: 8px; margin: 16px 0 0; }}
    .ma-summary div {{
      display: flex;
      justify-content: space-between;
      gap: 12px;
      border-top: 1px solid var(--line);
      padding-top: 8px;
    }}
    .technical-trigger {{ font-size: 13px; }}
    .decision-grid {{
      display: grid;
      grid-template-columns: minmax(0, 1.5fr) minmax(290px, .8fr);
      gap: 16px;
      margin-top: 16px;
    }}
    .section-heading {{
      display: flex;
      align-items: baseline;
      justify-content: space-between;
      gap: 12px;
      margin-bottom: 10px;
    }}
    .section-heading h2, .recent-panel > h2 {{ font-size: 19px; }}
    .section-heading span {{ color: var(--muted); font-size: 13px; }}
    .driver-head, .driver-row {{
      display: grid;
      grid-template-columns: 62px minmax(110px, 1fr) 100px 110px minmax(150px, 1.2fr) 70px;
      gap: 10px;
      align-items: center;
    }}
    .driver-head {{
      color: var(--muted);
      font-size: 12px;
      padding: 5px 0 7px;
      border-bottom: 1px solid var(--line);
    }}
    .driver-row {{ padding: 11px 0; border-top: 1px solid var(--line); font-size: 13px; }}
    .driver-row:first-child {{ border-top: 0; }}
    .driver-row small {{ display: block; color: var(--muted); font-size: 12px; }}
    .driver-state, .change-state, .quality {{
      display: inline-block;
      width: fit-content;
      border: 1px solid var(--line);
      border-radius: 999px;
      padding: 2px 8px;
      color: var(--muted);
      background: #f7f8f5;
      font-size: 12px;
      white-space: nowrap;
    }}
    .driver-state.state-supportive, .change-state.state-supportive {{ color: var(--support); background: #ecf6f0; }}
    .driver-state.state-neutral, .change-state.state-neutral {{ color: var(--neutral); background: #faf6e6; }}
    .driver-state.state-headwind, .change-state.state-headwind {{ color: var(--headwind); background: #fbeeee; }}
    .recent-panel > h2 {{ margin-bottom: 12px; }}
    .recent-change {{
      padding: 12px;
      border-left: 3px solid var(--neutral);
      background: #f7f8f5;
      border-radius: 9px;
    }}
    .recent-change + .recent-change {{ margin-top: 10px; }}
    .recent-change.tone-supportive {{ border-left-color: var(--support); }}
    .recent-change.tone-headwind {{ border-left-color: var(--headwind); }}
    .change-meta {{ display: flex; align-items: center; justify-content: space-between; gap: 8px; }}
    .change-meta small {{ color: var(--muted); font-size: 12px; }}
    .recent-change h3 {{ font-size: 16px; margin-top: 7px; }}
    .recent-change p {{ color: #334036; font-size: 13px; margin-top: 5px; }}
    .research-evidence {{ margin-top: 16px; padding: 0 16px 16px; }}
    .research-evidence > summary {{
      cursor: pointer;
      padding: 16px 0;
      display: flex;
      justify-content: space-between;
      gap: 16px;
    }}
    .research-evidence > summary span {{ color: var(--muted); font-size: 13px; text-align: right; }}
    .research-evidence[open] > summary {{ border-bottom: 1px solid var(--line); }}
    .research-evidence .wide {{ margin-top: 16px; }}
    .research-intro {{ color: #334036; margin-bottom: 12px; font-size: 13px; }}
    @media (max-width: 900px) {{
      .decision-header, .price-stage, .decision-grid, .chart-pair,
      .relationship-grid, .dual-axis-grid {{ grid-template-columns: 1fr; }}
      .posture {{ text-align: left; }}
      .driver-head {{ display: none; }}
      .driver-row {{ grid-template-columns: 54px minmax(120px, 1fr) minmax(90px, auto) minmax(90px, auto); }}
      .driver-row > span:nth-last-child(-n+2) {{ grid-column: 2 / -1; }}
      .chart-box {{ min-width: 0; }}
    }}
    @media (max-width: 560px) {{
      main {{ padding: 18px 12px 32px; }}
      h1 {{ font-size: 28px; }}
      .price-value {{ font-size: 32px; }}
      .driver-row {{ grid-template-columns: 48px minmax(0, 1fr); }}
      .driver-row > strong, .driver-row > span:nth-child(n+3) {{ grid-column: 2 / -1; }}
      .research-evidence {{ padding: 0 12px 12px; }}
      .research-evidence > summary {{ flex-direction: column; }}
      .research-evidence > summary span {{ text-align: left; }}
      .metric-strip {{ grid-template-columns: 1fr; }}
      .wide table {{ min-width: 680px; }}
    }}
  </style>
</head>
<body>
<main>
  <header class="decision-header">
    <div>
      <div class="kicker">黄金决策看板</div>
      <h1>{escape(dashboard['title'])}</h1>
      <p>{escape(updated)} · {escape(availability)}</p>
    </div>
    <aside class="posture state-{escape(dashboard['posture_state'])}">
      <span>驱动合成</span>
      <strong>{escape(dashboard['posture'])}</strong>
      <small>{escape(summary)}</small>
    </aside>
  </header>
  {price_section}
  <section class="decision-grid">
    {driver_section}
    {recent_changes_section}
  </section>
  {research_details}
</main>
<script>
{GOLD_CHART_SCRIPT}
</script>
</body>
</html>
"""
    return "\n".join(line.rstrip() for line in html.splitlines()) + "\n"


def main():
    SITE_DIR.mkdir(exist_ok=True)
    dashboard = read_dashboard_data()
    html = build_html(dashboard)
    OUTPUT_FILE.write_text(html, encoding="utf-8")
    print(f"Wrote {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
